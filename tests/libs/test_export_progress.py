"""
Tests for export progress tracking functionality.
"""

from datetime import datetime
from io import BytesIO
from unittest.mock import MagicMock, Mock, patch

import pytest
from django.core.cache import cache
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from libs.export_xlsx import ExportProgressTracker, XLSXGenerator, get_progress
from libs.export_xlsx.constants import (
    DEFAULT_PROGRESS_CHUNK_SIZE,
    REDIS_PROGRESS_EXPIRE_SECONDS,
    REDIS_PROGRESS_KEY_PREFIX,
)


class ExportProgressTrackerTests(TestCase):
    """Test cases for ExportProgressTracker."""

    def setUp(self):
        """Set up test fixtures."""
        self.task_id = "test-task-123"
        self.celery_task = MagicMock()
        self.tracker = ExportProgressTracker(task_id=self.task_id, celery_task=self.celery_task)
        # Clear any existing cache data
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    def test_set_total(self):
        """Test setting total rows."""
        self.tracker.set_total(1000)

        self.assertEqual(self.tracker.total_rows, 1000)
        self.assertEqual(self.tracker.processed_rows, 0)
        self.assertIsNotNone(self.tracker.start_time)

        # Verify Redis update
        progress = cache.get(self.tracker.redis_key)
        self.assertIsNotNone(progress)
        self.assertEqual(progress["total_rows"], 1000)
        self.assertEqual(progress["processed_rows"], 0)
        self.assertEqual(progress["percent"], 0)

    def test_update_progress(self):
        """Test updating progress."""
        self.tracker.set_total(1000)
        self.tracker.update(250)

        self.assertEqual(self.tracker.processed_rows, 250)

        # Verify Redis update
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["processed_rows"], 250)
        self.assertEqual(progress["percent"], 25)
        self.assertIn("speed_rows_per_sec", progress)
        self.assertIn("eta_seconds", progress)

    def test_multiple_updates(self):
        """Test multiple progress updates."""
        self.tracker.set_total(1000)
        self.tracker.update(100)
        self.tracker.update(100)
        self.tracker.update(100)

        self.assertEqual(self.tracker.processed_rows, 300)

        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["processed_rows"], 300)
        self.assertEqual(progress["percent"], 30)

    def test_set_completed(self):
        """Test marking export as completed."""
        self.tracker.set_total(1000)
        self.tracker.update(1000)
        self.tracker.set_completed(file_url="https://example.com/file.xlsx", file_path="exports/file.xlsx")

        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["status"], "SUCCESS")
        self.assertEqual(progress["percent"], 100)
        self.assertEqual(progress["file_url"], "https://example.com/file.xlsx")
        self.assertEqual(progress["file_path"], "exports/file.xlsx")

    def test_set_failed(self):
        """Test marking export as failed."""
        self.tracker.set_total(1000)
        self.tracker.update(500)
        self.tracker.set_failed("Test error message")

        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["status"], "FAILURE")
        self.assertEqual(progress["error"], "Test error message")
        self.assertEqual(progress["processed_rows"], 500)

    def test_progress_percentage_calculation(self):
        """Test progress percentage calculation."""
        self.tracker.set_total(1000)

        # 0%
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 0)

        # 25%
        self.tracker.update(250)
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 25)

        # 50%
        self.tracker.update(250)
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 50)

        # 100%
        self.tracker.update(500)
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 100)

    def test_celery_state_update(self):
        """Test that Celery task state is updated."""
        self.tracker.set_total(1000)
        self.tracker.update(250)

        # Verify update_state was called
        self.celery_task.update_state.assert_called()
        call_args = self.celery_task.update_state.call_args
        self.assertEqual(call_args[1]["state"], "PROGRESS")
        self.assertIn("percent", call_args[1]["meta"])
        self.assertIn("processed_rows", call_args[1]["meta"])

    def test_no_celery_task(self):
        """Test tracker works without Celery task."""
        tracker = ExportProgressTracker(task_id="test-task-456", celery_task=None)
        tracker.set_total(100)
        tracker.update(50)

        # Should still update Redis
        progress = cache.get(tracker.redis_key)
        self.assertIsNotNone(progress)
        self.assertEqual(progress["processed_rows"], 50)

    def test_redis_expiration(self):
        """Test Redis key has expiration set."""
        with patch("libs.export_xlsx.progress.cache.set") as mock_set:
            self.tracker.set_total(1000)

            mock_set.assert_called()
            call_args = mock_set.call_args
            self.assertEqual(call_args[1]["timeout"], REDIS_PROGRESS_EXPIRE_SECONDS)


class GetProgressTests(TestCase):
    """Test cases for get_progress utility function."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    def test_get_existing_progress(self):
        """Test retrieving existing progress."""
        task_id = "test-task-789"
        redis_key = f"{REDIS_PROGRESS_KEY_PREFIX}{task_id}"

        progress_data = {
            "status": "PROGRESS",
            "percent": 50,
            "processed_rows": 500,
            "total_rows": 1000,
        }
        cache.set(redis_key, progress_data, timeout=3600)

        result = get_progress(task_id)
        self.assertIsNotNone(result)
        self.assertEqual(result["percent"], 50)
        self.assertEqual(result["processed_rows"], 500)

    def test_get_nonexistent_progress(self):
        """Test retrieving non-existent progress."""
        result = get_progress("nonexistent-task-id")
        self.assertIsNone(result)


class XLSXGeneratorProgressTests(TestCase):
    """Test cases for XLSXGenerator progress callbacks."""

    def setUp(self):
        """Set up test fixtures."""
        self.progress_updates = []

        def progress_callback(rows_processed: int):
            self.progress_updates.append(rows_processed)

        self.progress_callback = progress_callback

    def test_generator_without_callback(self):
        """Test generator works without progress callback."""
        generator = XLSXGenerator()
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name", "Age"],
                    "field_names": ["name", "age"],
                    "data": [{"name": f"User {i}", "age": 20 + i} for i in range(100)],
                }
            ]
        }

        file_content = generator.generate(schema)
        self.assertIsInstance(file_content, BytesIO)

        # Verify data
        wb = load_workbook(file_content)
        ws = wb["Test"]
        self.assertEqual(ws.cell(2, 1).value, "User 0")
        self.assertEqual(ws.cell(101, 1).value, "User 99")

    def test_generator_with_callback_small_dataset(self):
        """Test progress callback with small dataset (< chunk_size)."""
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=500)
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(100)],
                }
            ]
        }

        generator.generate(schema)

        # Should have one update at the end for remaining rows
        self.assertEqual(len(self.progress_updates), 1)
        self.assertEqual(sum(self.progress_updates), 100)

    def test_generator_with_callback_large_dataset(self):
        """Test progress callback with large dataset (> chunk_size)."""
        chunk_size = 500
        total_rows = 2500
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=chunk_size)
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(total_rows)],
                }
            ]
        }

        generator.generate(schema)

        # Should have 5 updates (500 * 5 = 2500)
        self.assertEqual(len(self.progress_updates), 5)
        self.assertEqual(sum(self.progress_updates), total_rows)

    def test_generator_with_callback_multiple_sheets(self):
        """Test progress callback with multiple sheets."""
        chunk_size = 500
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=chunk_size)
        schema = {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(600)],
                },
                {
                    "name": "Sheet2",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(700)],
                },
            ]
        }

        generator.generate(schema)

        # Total: 1300 rows
        # Updates: 500, 500, 300 (3 updates)
        self.assertEqual(len(self.progress_updates), 3)
        self.assertEqual(sum(self.progress_updates), 1300)

    def test_generator_callback_frequency(self):
        """Test progress callback is called at correct intervals."""
        chunk_size = 100
        total_rows = 350
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=chunk_size)
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(total_rows)],
                }
            ]
        }

        generator.generate(schema)

        # Updates should be: 100, 100, 100, 50
        self.assertEqual(len(self.progress_updates), 4)
        self.assertEqual(self.progress_updates[0], 100)
        self.assertEqual(self.progress_updates[1], 100)
        self.assertEqual(self.progress_updates[2], 100)
        self.assertEqual(self.progress_updates[3], 50)

    def test_calculate_total_rows(self):
        """Test total rows calculation."""
        generator = XLSXGenerator()
        schema = {
            "sheets": [
                {"name": "Sheet1", "data": [{"name": "User 1"}, {"name": "User 2"}]},
                {"name": "Sheet2", "data": [{"name": "User 3"}, {"name": "User 4"}, {"name": "User 5"}]},
            ]
        }

        total = generator._calculate_total_rows(schema)
        self.assertEqual(total, 5)

    def test_calculate_total_rows_empty_sheets(self):
        """Test total rows calculation with empty sheets."""
        generator = XLSXGenerator()
        schema = {
            "sheets": [
                {"name": "Sheet1", "data": []},
                {"name": "Sheet2", "data": [{"name": "User 1"}]},
            ]
        }

        total = generator._calculate_total_rows(schema)
        self.assertEqual(total, 1)


@override_settings(
    EXPORTER_CELERY_ENABLED=True,
    EXPORTER_STORAGE_BACKEND="local",
    EXPORTER_PROGRESS_CHUNK_SIZE=100,
)
class GenerateXLSXTaskProgressTests(TestCase):
    """Test cases for generate_xlsx_task with progress tracking."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    def test_task_publishes_progress(self, mock_storage_backend):
        """Test that task publishes progress during execution."""
        # Mock storage
        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        # Import task
        from libs.export_xlsx.tasks import generate_xlsx_task

        # Create schema with 250 rows (should trigger progress updates)
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(250)],
                }
            ]
        }

        # Create a mock task with request
        task = generate_xlsx_task
        task.request = Mock()
        task.request.id = "test-task-progress-123"

        # Execute task
        result = task(schema, filename="test.xlsx", storage_backend="local")

        # Verify result
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["file_url"], "https://example.com/test.xlsx")

        # Verify progress was published to Redis
        progress = get_progress(task.request.id)
        self.assertIsNotNone(progress)
        self.assertEqual(progress["status"], "SUCCESS")
        self.assertEqual(progress["percent"], 100)
        self.assertEqual(progress["total_rows"], 250)
        self.assertEqual(progress["processed_rows"], 250)

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    def test_task_handles_failure(self, mock_storage_backend):
        """Test that task handles failures correctly."""
        # Make storage raise an error
        mock_storage = Mock()
        mock_storage.save.side_effect = Exception("Storage error")
        mock_storage_backend.return_value = mock_storage

        from libs.export_xlsx.tasks import generate_xlsx_task

        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": "User 1"}],
                }
            ]
        }

        task = generate_xlsx_task
        task.request = Mock()
        task.request.id = "test-task-failure-456"

        result = task(schema, filename="test.xlsx", storage_backend="local")

        # Verify result
        self.assertEqual(result["status"], "error")
        self.assertIn("Storage error", result["error"])

        # Verify progress shows failure
        progress = get_progress(task.request.id)
        self.assertIsNotNone(progress)
        self.assertEqual(progress["status"], "FAILURE")
        self.assertIn("Storage error", progress["error"])
