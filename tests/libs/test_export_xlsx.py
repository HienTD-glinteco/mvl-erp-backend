"""
Tests for XLSX export module.
"""

from io import BytesIO
from unittest.mock import MagicMock, patch

from django.contrib.auth import get_user_model
from django.db import models
from django.test import TestCase, override_settings
from openpyxl import load_workbook

from libs.export_xlsx import SchemaBuilder, XLSXGenerator, get_storage_backend
from libs.export_xlsx.constants import ERROR_INVALID_SCHEMA, STORAGE_LOCAL, STORAGE_S3

User = get_user_model()


class SchemaBuilderTests(TestCase):
    """Test cases for SchemaBuilder."""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        class SchemaBuilderTestsModel(models.Model):
            """Test model for schema builder tests."""

            name = models.CharField(max_length=100, verbose_name="Name")
            email = models.EmailField(verbose_name="Email Address")
            age = models.IntegerField(verbose_name="Age")
            is_active = models.BooleanField(default=True, verbose_name="Active")
            created_at = models.DateTimeField(auto_now_add=True)

            class Meta:
                app_label = "test"
                verbose_name = "Test Item"
                verbose_name_plural = "Test Items"

        cls.TestModel = SchemaBuilderTestsModel

    def setUp(self):
        """Set up test fixtures."""
        self.builder = SchemaBuilder()

    def test_build_from_model_without_data(self):
        """Test building schema from model without data."""
        schema = self.builder.build_from_model(self.TestModel)

        self.assertIn("sheets", schema)
        self.assertEqual(len(schema["sheets"]), 1)

        sheet = schema["sheets"][0]
        self.assertEqual(sheet["name"], "Test Items")
        self.assertIn("headers", sheet)
        self.assertIn("field_names", sheet)
        self.assertEqual(sheet["data"], [])

        # Check that excluded fields are not present
        self.assertNotIn("id", sheet["field_names"])
        self.assertNotIn("created_at", sheet["field_names"])
        self.assertNotIn("updated_at", sheet["field_names"])

    def test_build_from_model_with_queryset(self):
        """Test building schema from model with queryset data."""
        # Create mock queryset
        mock_obj = MagicMock()
        mock_obj.name = "John Doe"
        mock_obj.email = "john@example.com"
        mock_obj.age = 30
        mock_obj.is_active = True

        mock_queryset = [mock_obj]

        schema = self.builder.build_from_model(self.TestModel, mock_queryset)

        sheet = schema["sheets"][0]
        self.assertEqual(len(sheet["data"]), 1)

        # Check data values
        row = sheet["data"][0]
        self.assertEqual(row["name"], "John Doe")
        self.assertEqual(row["email"], "john@example.com")
        self.assertEqual(row["age"], 30)
        self.assertEqual(row["is_active"], "Yes")

    def test_get_field_label(self):
        """Test field label generation."""
        field = self.TestModel._meta.get_field("name")
        label = self.builder._get_field_label(field)
        self.assertEqual(label, "Name")

    def test_excluded_fields(self):
        """Test that excluded fields are not included."""
        custom_excluded = {"name", "email"}
        builder = SchemaBuilder(excluded_fields=custom_excluded)

        schema = builder.build_from_model(self.TestModel)
        sheet = schema["sheets"][0]

        self.assertNotIn("name", sheet["field_names"])
        self.assertNotIn("email", sheet["field_names"])
        self.assertIn("age", sheet["field_names"])


class XLSXGeneratorTests(TestCase):
    """Test cases for XLSXGenerator."""

    def setUp(self):
        """Set up test fixtures."""
        self.generator = XLSXGenerator()

    def test_generate_simple_schema(self):
        """Test generating XLSX from simple schema."""
        schema = {
            "sheets": [
                {
                    "name": "Test Sheet",
                    "headers": ["Name", "Email", "Age"],
                    "field_names": ["name", "email", "age"],
                    "data": [
                        {"name": "John Doe", "email": "john@example.com", "age": 30},
                        {"name": "Jane Smith", "email": "jane@example.com", "age": 25},
                    ],
                }
            ]
        }

        file_content = self.generator.generate(schema)

        # Verify it's a valid BytesIO object
        self.assertIsInstance(file_content, BytesIO)

        # Load and verify the workbook
        wb = load_workbook(file_content)
        self.assertIn("Test Sheet", wb.sheetnames)

        ws = wb["Test Sheet"]
        # Check headers
        self.assertEqual(ws.cell(1, 1).value, "Name")
        self.assertEqual(ws.cell(1, 2).value, "Email")
        self.assertEqual(ws.cell(1, 3).value, "Age")

        # Check data
        self.assertEqual(ws.cell(2, 1).value, "John Doe")
        self.assertEqual(ws.cell(2, 2).value, "john@example.com")
        self.assertEqual(ws.cell(2, 3).value, 30)

    def test_generate_with_groups(self):
        """Test generating XLSX with grouped headers."""
        schema = {
            "sheets": [
                {
                    "name": "Grouped Sheet",
                    "headers": ["Name", "Email", "Age", "Status"],
                    "field_names": ["name", "email", "age", "status"],
                    "groups": [
                        {"title": "Personal Info", "span": 2},
                        {"title": "Details", "span": 2},
                    ],
                    "data": [
                        {"name": "John", "email": "john@example.com", "age": 30, "status": "Active"},
                    ],
                }
            ]
        }

        file_content = self.generator.generate(schema)

        wb = load_workbook(file_content)
        ws = wb["Grouped Sheet"]

        # Check group headers
        self.assertEqual(ws.cell(1, 1).value, "Personal Info")
        self.assertEqual(ws.cell(1, 3).value, "Details")

        # Check column headers
        self.assertEqual(ws.cell(2, 1).value, "Name")
        self.assertEqual(ws.cell(2, 2).value, "Email")

    def test_generate_with_merge_rules(self):
        """Test generating XLSX with merge rules."""
        schema = {
            "sheets": [
                {
                    "name": "Merged Sheet",
                    "headers": ["Project", "Task", "Hours"],
                    "field_names": ["project", "task", "hours"],
                    "merge_rules": ["project"],
                    "data": [
                        {"project": "Project A", "task": "Task 1", "hours": 10},
                        {"project": "Project A", "task": "Task 2", "hours": 15},
                        {"project": "Project B", "task": "Task 3", "hours": 20},
                    ],
                }
            ]
        }

        file_content = self.generator.generate(schema)

        wb = load_workbook(file_content)
        ws = wb["Merged Sheet"]

        # Check data values for merged cells
        self.assertEqual(ws.cell(2, 1).value, "Project A")
        self.assertIsNone(ws.cell(3, 1).value)  # Merged cell, should be None
        self.assertEqual(ws.cell(4, 1).value, "Project B")

    def test_generate_invalid_schema(self):
        """Test that invalid schema raises error."""
        with self.assertRaises(ValueError) as context:
            self.generator.generate({})

        self.assertIn(ERROR_INVALID_SCHEMA, str(context.exception))

    def test_generate_multiple_sheets(self):
        """Test generating XLSX with multiple sheets."""
        schema = {
            "sheets": [
                {
                    "name": "Sheet 1",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": "John"}],
                },
                {
                    "name": "Sheet 2",
                    "headers": ["Email"],
                    "field_names": ["email"],
                    "data": [{"email": "john@example.com"}],
                },
            ]
        }

        file_content = self.generator.generate(schema)

        wb = load_workbook(file_content)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn("Sheet 1", wb.sheetnames)
        self.assertIn("Sheet 2", wb.sheetnames)


@override_settings(
    EXPORTER_STORAGE_BACKEND=STORAGE_LOCAL,
    EXPORTER_LOCAL_STORAGE_PATH="test_exports",
)
class StorageBackendTests(TestCase):
    """Test cases for storage backends."""

    def test_get_local_storage_backend(self):
        """Test getting local storage backend."""
        backend = get_storage_backend(STORAGE_LOCAL)
        self.assertIsNotNone(backend)
        self.assertEqual(backend.storage_path, "test_exports")
        # Verify it uses FileSystemStorage, not default_storage
        from django.core.files.storage import FileSystemStorage

        self.assertIsInstance(backend.storage, FileSystemStorage)

    # def test_get_s3_storage_backend(self):
    #     """Test getting S3 storage backend."""
    #     backend = get_storage_backend(STORAGE_S3)
    #     self.assertIsNotNone(backend)

    def test_invalid_storage_backend(self):
        """Test that invalid storage backend raises error."""
        with self.assertRaises(ValueError):
            get_storage_backend("invalid")

    @patch("libs.export_xlsx.storage.FileSystemStorage.save")
    @patch("libs.export_xlsx.storage.FileSystemStorage.url")
    def test_local_storage_save(self, mock_url, mock_save):
        """Test saving file to local storage using FileSystemStorage."""
        mock_save.return_value = "20250101_120000_test.xlsx"
        mock_url.return_value = "/media/test_exports/20250101_120000_test.xlsx"

        backend = get_storage_backend(STORAGE_LOCAL)
        file_content = BytesIO(b"test content")

        file_path = backend.save(file_content, "test.xlsx")
        file_url = backend.get_url(file_path)

        self.assertTrue(mock_save.called)
        self.assertIsNotNone(file_path)
        self.assertIsNotNone(file_url)
