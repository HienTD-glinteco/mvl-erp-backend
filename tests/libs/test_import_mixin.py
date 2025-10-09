"""
Tests for ImportXLSXMixin.
"""

import io

import pytest
from django.db import models
from django.test import TestCase
from openpyxl import Workbook
from rest_framework import serializers
from rest_framework.test import APIRequestFactory

from libs import ImportXLSXMixin
from libs.base_viewset import BaseModelViewSet


# Test fixtures - Mock model and ViewSet
class TestImportModel(models.Model):
    """Mock model for testing imports"""

    name = models.CharField(max_length=100)
    email = models.EmailField()
    age = models.IntegerField(null=True, blank=True)

    class Meta:
        app_label = "test_import"


class TestImportSerializer(serializers.ModelSerializer):
    """Serializer for test model"""

    class Meta:
        model = TestImportModel
        fields = ["name", "email", "age"]


class TestImportViewSet(ImportXLSXMixin, BaseModelViewSet):
    """Test ViewSet with import functionality"""

    queryset = TestImportModel.objects.all()
    serializer_class = TestImportSerializer
    module = "Test"
    submodule = "Import Test"
    permission_prefix = "test_import"


class TestImportViewSetWithCustomSchema(ImportXLSXMixin, BaseModelViewSet):
    """Test ViewSet with custom import schema"""

    queryset = TestImportModel.objects.all()
    serializer_class = TestImportSerializer
    module = "Test"
    submodule = "Import Test"
    permission_prefix = "test_import_custom"

    def get_import_schema(self, request, file):
        return {
            "fields": ["name", "email"],
            "required": ["name"],
        }


class ImportXLSXMixinTestCase(TestCase):
    """Test ImportXLSXMixin functionality"""

    def setUp(self):
        """Set up test fixtures"""
        self.factory = APIRequestFactory()
        self.viewset = TestImportViewSet()

    def _create_xlsx_file(self, data, headers=None):
        """
        Helper to create XLSX file in memory.

        Args:
            data: List of rows (list of lists)
            headers: Optional list of header names

        Returns:
            File-like object containing XLSX data
        """
        workbook = Workbook()
        sheet = workbook.active

        # Add headers
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_idx, value=header)
            start_row = 2
        else:
            start_row = 1

        # Add data rows
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        output.name = "test.xlsx"
        return output

    def test_auto_generate_schema(self):
        """Test auto schema generation from model fields"""
        # Arrange
        viewset = TestImportViewSet()

        # Act
        schema = viewset._auto_generate_schema()

        # Assert
        self.assertIn("fields", schema)
        self.assertIn("required", schema)
        self.assertIn("name", schema["fields"])
        self.assertIn("email", schema["fields"])
        self.assertIn("age", schema["fields"])
        # ID and timestamps should be excluded
        self.assertNotIn("id", schema["fields"])

    def test_auto_generate_schema_required_fields(self):
        """Test auto schema identifies required fields"""
        # Arrange
        viewset = TestImportViewSet()

        # Act
        schema = viewset._auto_generate_schema()

        # Assert
        # name and email are required (blank=False, null=False)
        self.assertIn("name", schema["required"])
        self.assertIn("email", schema["required"])
        # age is optional (null=True, blank=True)
        self.assertNotIn("age", schema["required"])

    def test_custom_import_schema(self):
        """Test custom import schema override"""
        # Arrange
        viewset = TestImportViewSetWithCustomSchema()
        request = self.factory.get("/")
        file = None

        # Act
        schema = viewset.get_import_schema(request, file)

        # Assert
        self.assertEqual(schema["fields"], ["name", "email"])
        self.assertEqual(schema["required"], ["name"])

    def test_map_headers_to_fields(self):
        """Test header to field mapping"""
        # Arrange
        viewset = TestImportViewSet()
        headers = ["Name", "Email", "Age"]
        fields = ["name", "email", "age"]

        # Act
        mapping = viewset._map_headers_to_fields(headers, fields)

        # Assert
        self.assertEqual(mapping["Name"], "name")
        self.assertEqual(mapping["Email"], "email")
        self.assertEqual(mapping["Age"], "age")

    def test_map_headers_case_insensitive(self):
        """Test header mapping is case-insensitive"""
        # Arrange
        viewset = TestImportViewSet()
        headers = ["NAME", "email", "AgE"]
        fields = ["name", "email", "age"]

        # Act
        mapping = viewset._map_headers_to_fields(headers, fields)

        # Assert
        self.assertEqual(mapping["NAME"], "name")
        self.assertEqual(mapping["email"], "email")
        self.assertEqual(mapping["AgE"], "age")

    def test_map_headers_with_underscores(self):
        """Test header mapping handles spaces and underscores"""
        # Arrange
        viewset = TestImportViewSet()
        headers = ["User Name", "Email Address"]
        fields = ["user_name", "email_address"]

        # Act
        mapping = viewset._map_headers_to_fields(headers, fields)

        # Assert
        self.assertEqual(mapping["User Name"], "user_name")
        self.assertEqual(mapping["Email Address"], "email_address")

    def test_parse_xlsx_file_valid_data(self):
        """Test parsing valid XLSX file"""
        # Arrange
        viewset = TestImportViewSet()
        schema = {"fields": ["name", "email", "age"], "required": ["name", "email"]}
        headers = ["name", "email", "age"]
        data = [
            ["John Doe", "john@example.com", 30],
            ["Jane Smith", "jane@example.com", 25],
        ]
        file = self._create_xlsx_file(data, headers)

        # Act
        parsed_data, errors = viewset._parse_xlsx_file(file, schema)

        # Assert
        self.assertEqual(len(parsed_data), 2)
        self.assertEqual(len(errors), 0)
        self.assertEqual(parsed_data[0]["name"], "John Doe")
        self.assertEqual(parsed_data[1]["name"], "Jane Smith")

    def test_parse_xlsx_file_missing_required_field(self):
        """Test parsing XLSX with missing required field"""
        # Arrange
        viewset = TestImportViewSet()
        schema = {"fields": ["name", "email"], "required": ["name", "email"]}
        headers = ["name", "email"]
        data = [
            ["John Doe", "john@example.com"],
            ["Jane Smith", ""],  # Missing email
        ]
        file = self._create_xlsx_file(data, headers)

        # Act
        parsed_data, errors = viewset._parse_xlsx_file(file, schema)

        # Assert
        self.assertEqual(len(errors), 1)
        self.assertEqual(errors[0]["row"], 3)
        self.assertIn("email", errors[0]["errors"])

    def test_parse_xlsx_file_invalid_email(self):
        """Test parsing XLSX with invalid email format"""
        # Arrange
        viewset = TestImportViewSet()
        schema = {"fields": ["name", "email"], "required": ["name", "email"]}
        headers = ["name", "email"]
        data = [
            ["John Doe", "invalid-email"],  # Invalid email
        ]
        file = self._create_xlsx_file(data, headers)

        # Act
        parsed_data, errors = viewset._parse_xlsx_file(file, schema)

        # Assert
        self.assertEqual(len(errors), 1)
        self.assertIn("email", errors[0]["errors"])

    def test_parse_xlsx_file_skip_empty_rows(self):
        """Test parsing XLSX skips empty rows"""
        # Arrange
        viewset = TestImportViewSet()
        schema = {"fields": ["name", "email"], "required": ["name", "email"]}
        headers = ["name", "email"]
        data = [
            ["John Doe", "john@example.com"],
            ["", ""],  # Empty row
            ["Jane Smith", "jane@example.com"],
        ]
        file = self._create_xlsx_file(data, headers)

        # Act
        parsed_data, errors = viewset._parse_xlsx_file(file, schema)

        # Assert
        # Should parse 2 valid rows, skip 1 empty row
        self.assertEqual(len(parsed_data), 2)

    @pytest.mark.django_db
    def test_import_action_no_file(self):
        """Test import action returns error when no file provided"""
        # Arrange
        viewset = TestImportViewSet()
        request = self.factory.post("/import/")
        viewset.request = request

        # Act
        response = viewset.import_data(request)

        # Assert
        self.assertEqual(response.status_code, 400)
        self.assertIn("detail", response.data)

    @pytest.mark.django_db
    def test_import_action_invalid_file_type(self):
        """Test import action returns error for non-XLSX file"""
        # Arrange
        viewset = TestImportViewSet()
        file = io.BytesIO(b"not an xlsx file")
        file.name = "test.txt"
        request = self.factory.post("/import/", {"file": file})
        viewset.request = request

        # Act
        response = viewset.import_data(request)

        # Assert
        self.assertEqual(response.status_code, 400)
        self.assertIn("detail", response.data)

    @pytest.mark.django_db
    def test_import_action_empty_file(self):
        """Test import action handles empty file"""
        # Arrange
        viewset = TestImportViewSet()
        file = self._create_xlsx_file([], ["name", "email"])
        request = self.factory.post("/import/", {"file": file}, format="multipart")
        viewset.request = request

        # Act
        response = viewset.import_data(request)

        # Assert
        self.assertEqual(response.status_code, 400)


class ImportPermissionTestCase(TestCase):
    """Test import action generates correct permissions"""

    def test_import_action_registered_as_custom_action(self):
        """Test that import_data is registered as a custom action"""
        # Arrange & Act
        custom_actions = TestImportViewSet.get_custom_actions()

        # Assert
        self.assertIn("import_data", custom_actions)

    def test_import_permission_generated(self):
        """Test that import action generates permission metadata"""
        # Arrange & Act
        permissions = TestImportViewSet.get_registered_permissions()

        # Assert
        import_perm = next((p for p in permissions if p["code"] == "test_import.import_data"), None)
        self.assertIsNotNone(import_perm)
        self.assertEqual(import_perm["module"], "Test")
        self.assertEqual(import_perm["submodule"], "Import Test")
