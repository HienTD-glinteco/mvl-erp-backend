"""
Tests for export progress tracking functionality.
"""

from io import BytesIO
from unittest.mock import MagicMock, Mock, patch

from django.core.cache import cache
from django.test import SimpleTestCase, TestCase, override_settings
from openpyxl import load_workbook

from libs.export_xlsx import ExportProgressTracker, XLSXGenerator, get_progress
from libs.export_xlsx.constants import (
    REDIS_PROGRESS_EXPIRE_SECONDS,
    REDIS_PROGRESS_KEY_PREFIX,
)


@override_settings(
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    }
)
class ExportProgressTrackerTests(SimpleTestCase):
    """Test cases for ExportProgressTracker."""

    def setUp(self):
        """Set up test fixtures."""
        self.task_id = "test-task-123"
        self.celery_task = MagicMock()
        # Clear cache before creating tracker
        cache.clear()
        self.tracker = ExportProgressTracker(task_id=self.task_id, celery_task=self.celery_task)

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    def test_set_total(self):
        """Test setting total rows."""
        self.tracker.set_total(1000)

        self.assertEqual(self.tracker.total_rows, 1000)
        self.assertEqual(self.tracker.processed_rows, 0)
        self.assertIsNotNone(self.tracker.start_time)

        # Verify Redis update
        progress = cache.get(self.tracker.redis_key)
        self.assertIsNotNone(progress)
        self.assertEqual(progress["total_rows"], 1000)
        self.assertEqual(progress["processed_rows"], 0)
        self.assertEqual(progress["percent"], 0)

    def test_update_progress(self):
        """Test updating progress."""
        self.tracker.set_total(1000)
        self.tracker.update(250)

        self.assertEqual(self.tracker.processed_rows, 250)

        # Verify Redis update
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["processed_rows"], 250)
        self.assertEqual(progress["percent"], 25)
        self.assertIn("speed_rows_per_sec", progress)
        self.assertIn("eta_seconds", progress)

    def test_multiple_updates(self):
        """Test multiple progress updates."""
        self.tracker.set_total(1000)
        self.tracker.update(100)
        self.tracker.update(100)
        self.tracker.update(100)

        self.assertEqual(self.tracker.processed_rows, 300)

        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["processed_rows"], 300)
        self.assertEqual(progress["percent"], 30)

    def test_set_completed(self):
        """Test marking export as completed."""
        self.tracker.set_total(1000)
        self.tracker.update(1000)
        self.tracker.set_completed(file_url="https://example.com/file.xlsx", file_path="exports/file.xlsx")

        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["status"], "SUCCESS")
        self.assertEqual(progress["percent"], 100)
        self.assertEqual(progress["file_url"], "https://example.com/file.xlsx")
        self.assertEqual(progress["file_path"], "exports/file.xlsx")

    def test_set_failed(self):
        """Test marking export as failed."""
        self.tracker.set_total(1000)
        self.tracker.update(500)
        self.tracker.set_failed("Test error message")

        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["status"], "FAILURE")
        self.assertEqual(progress["error"], "Test error message")
        self.assertEqual(progress["processed_rows"], 500)

    def test_progress_percentage_calculation(self):
        """Test progress percentage calculation."""
        self.tracker.set_total(1000)

        # 0%
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 0)

        # 25%
        self.tracker.update(250)
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 25)

        # 50%
        self.tracker.update(250)
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 50)

        # 100%
        self.tracker.update(500)
        progress = cache.get(self.tracker.redis_key)
        self.assertEqual(progress["percent"], 100)

    def test_celery_state_update(self):
        """Test that Celery task state is updated."""
        self.tracker.set_total(1000)
        self.tracker.update(250)

        # Verify update_state was called
        self.celery_task.update_state.assert_called()
        call_args = self.celery_task.update_state.call_args
        self.assertEqual(call_args[1]["state"], "PROGRESS")
        self.assertIn("percent", call_args[1]["meta"])
        self.assertIn("processed_rows", call_args[1]["meta"])

    def test_no_celery_task(self):
        """Test tracker works without Celery task."""
        tracker = ExportProgressTracker(task_id="test-task-456", celery_task=None)
        tracker.set_total(100)
        tracker.update(50)

        # Should still update Redis
        progress = cache.get(tracker.redis_key)
        self.assertIsNotNone(progress)
        self.assertEqual(progress["processed_rows"], 50)

    def test_redis_expiration(self):
        """Test Redis key has expiration set."""
        with patch("libs.export_xlsx.progress.cache.set") as mock_set:
            self.tracker.set_total(1000)

            mock_set.assert_called()
            call_args = mock_set.call_args
            self.assertEqual(call_args[1]["timeout"], REDIS_PROGRESS_EXPIRE_SECONDS)


@override_settings(
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    }
)
class GetProgressTests(SimpleTestCase):
    """Test cases for get_progress utility function."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    def test_get_existing_progress(self):
        """Test retrieving existing progress."""
        task_id = "test-task-789"
        redis_key = f"{REDIS_PROGRESS_KEY_PREFIX}{task_id}"

        progress_data = {
            "status": "PROGRESS",
            "percent": 50,
            "processed_rows": 500,
            "total_rows": 1000,
        }
        cache.set(redis_key, progress_data, timeout=3600)

        result = get_progress(task_id)
        self.assertIsNotNone(result)
        self.assertEqual(result["percent"], 50)
        self.assertEqual(result["processed_rows"], 500)

    def test_get_nonexistent_progress(self):
        """Test retrieving non-existent progress."""
        result = get_progress("nonexistent-task-id")
        self.assertIsNone(result)


@override_settings(
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    }
)
class XLSXGeneratorProgressTests(SimpleTestCase):
    """Test cases for XLSXGenerator progress callbacks."""

    def setUp(self):
        """Set up test fixtures."""
        self.progress_updates = []

        def progress_callback(rows_processed: int):
            self.progress_updates.append(rows_processed)

        self.progress_callback = progress_callback

    def test_generator_without_callback(self):
        """Test generator works without progress callback."""
        generator = XLSXGenerator()
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name", "Age"],
                    "field_names": ["name", "age"],
                    "data": [{"name": f"User {i}", "age": 20 + i} for i in range(100)],
                }
            ]
        }

        file_content = generator.generate(schema)
        self.assertIsInstance(file_content, BytesIO)

        # Verify data
        wb = load_workbook(file_content)
        ws = wb["Test"]
        self.assertEqual(ws.cell(2, 1).value, "User 0")
        self.assertEqual(ws.cell(101, 1).value, "User 99")

    def test_generator_with_callback_small_dataset(self):
        """Test progress callback with small dataset (< chunk_size)."""
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=500)
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(100)],
                }
            ]
        }

        generator.generate(schema)

        # Should have one update at the end for remaining rows
        self.assertEqual(len(self.progress_updates), 1)
        self.assertEqual(sum(self.progress_updates), 100)

    def test_generator_with_callback_large_dataset(self):
        """Test progress callback with large dataset (> chunk_size)."""
        chunk_size = 500
        total_rows = 2500
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=chunk_size)
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(total_rows)],
                }
            ]
        }

        generator.generate(schema)

        # Should have 5 updates (500 * 5 = 2500)
        self.assertEqual(len(self.progress_updates), 5)
        self.assertEqual(sum(self.progress_updates), total_rows)

    def test_generator_with_callback_multiple_sheets(self):
        """Test progress callback with multiple sheets."""
        chunk_size = 500
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=chunk_size)
        schema = {
            "sheets": [
                {
                    "name": "Sheet1",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(600)],
                },
                {
                    "name": "Sheet2",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(700)],
                },
            ]
        }

        generator.generate(schema)

        # Total: 1300 rows
        # Updates: 500 (sheet1), 100 (sheet1 remaining), 500 (sheet2), 200 (sheet2 remaining) = 4 updates
        self.assertEqual(len(self.progress_updates), 4)
        self.assertEqual(sum(self.progress_updates), 1300)

    def test_generator_callback_frequency(self):
        """Test progress callback is called at correct intervals."""
        chunk_size = 100
        total_rows = 350
        generator = XLSXGenerator(progress_callback=self.progress_callback, chunk_size=chunk_size)
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(total_rows)],
                }
            ]
        }

        generator.generate(schema)

        # Updates should be: 100, 100, 100, 50
        self.assertEqual(len(self.progress_updates), 4)
        self.assertEqual(self.progress_updates[0], 100)
        self.assertEqual(self.progress_updates[1], 100)
        self.assertEqual(self.progress_updates[2], 100)
        self.assertEqual(self.progress_updates[3], 50)

    def test_calculate_total_rows(self):
        """Test total rows calculation."""
        generator = XLSXGenerator()
        schema = {
            "sheets": [
                {"name": "Sheet1", "data": [{"name": "User 1"}, {"name": "User 2"}]},
                {"name": "Sheet2", "data": [{"name": "User 3"}, {"name": "User 4"}, {"name": "User 5"}]},
            ]
        }

        total = generator._calculate_total_rows(schema)
        self.assertEqual(total, 5)

    def test_calculate_total_rows_empty_sheets(self):
        """Test total rows calculation with empty sheets."""
        generator = XLSXGenerator()
        schema = {
            "sheets": [
                {"name": "Sheet1", "data": []},
                {"name": "Sheet2", "data": [{"name": "User 1"}]},
            ]
        }

        total = generator._calculate_total_rows(schema)
        self.assertEqual(total, 1)


@override_settings(
    EXPORTER_CELERY_ENABLED=True,
    EXPORTER_STORAGE_BACKEND="local",
    EXPORTER_PROGRESS_CHUNK_SIZE=100,
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    },
)
class GenerateXLSXTaskProgressTests(SimpleTestCase):
    """Test cases for generate_xlsx_task with progress tracking."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    @patch("libs.export_xlsx.tasks.XLSXGenerator")
    def test_task_publishes_progress(self, mock_generator_class, mock_tracker_class, mock_storage_backend):
        """Test that task publishes progress during execution."""
        # Mock XLSXGenerator
        mock_generator = Mock()
        mock_file_content = Mock()
        mock_generator.generate.return_value = mock_file_content
        mock_generator_class.return_value = mock_generator

        # Mock storage
        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        # Mock progress tracker
        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        # Import the task module to get the actual function (not bound)
        from libs.export_xlsx import tasks

        # Create schema with 250 rows
        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": f"User {i}"} for i in range(250)],
                }
            ]
        }

        # Create a mock self for the bound task
        mock_self = Mock()
        mock_self.request = Mock()
        mock_self.request.id = "test-task-123"

        # Call the underlying function directly
        result = tasks.generate_xlsx_task.run(schema, filename="test.xlsx", storage_backend="local")

        # Verify result
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["file_url"], "https://example.com/test.xlsx")

        # Verify progress tracker was initialized
        self.assertTrue(mock_tracker_class.called)

        # Verify progress methods were called
        mock_tracker.set_total.assert_called_once_with(250)
        mock_tracker.set_completed.assert_called_once()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    @patch("libs.export_xlsx.tasks.XLSXGenerator")
    def test_task_handles_failure(self, mock_generator_class, mock_tracker_class, mock_storage_backend):
        """Test that task handles failures correctly."""
        # Mock XLSXGenerator
        mock_generator = Mock()
        mock_file_content = Mock()
        mock_generator.generate.return_value = mock_file_content
        mock_generator_class.return_value = mock_generator

        # Make storage raise an error
        mock_storage = Mock()
        mock_storage.save.side_effect = Exception("Storage error")
        mock_storage_backend.return_value = mock_storage

        # Mock progress tracker
        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from libs.export_xlsx import tasks

        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": "User 1"}],
                }
            ]
        }

        result = tasks.generate_xlsx_task.run(schema, filename="test.xlsx", storage_backend="local")

        # Verify result
        self.assertEqual(result["status"], "error")
        self.assertIn("Storage error", result["error"])

        # Verify progress tracker recorded failure
        mock_tracker.set_failed.assert_called_once()
        error_call = mock_tracker.set_failed.call_args[0][0]
        self.assertIn("Storage error", error_call)


@override_settings(
    EXPORTER_CELERY_ENABLED=True,
    EXPORTER_STORAGE_BACKEND="local",
    EXPORTER_PROGRESS_CHUNK_SIZE=100,
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    },
)
class GenerateXLSXFromQuerysetTaskTests(TestCase):
    """Test cases for generate_xlsx_from_queryset_task."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    def test_task_builds_schema_from_queryset(self, mock_tracker_class, mock_storage_backend):
        """Test that task builds schema from queryset parameters."""
        # Mock storage
        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        # Mock progress tracker
        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from apps.core.models import Role
        from libs.export_xlsx import tasks

        # Create test data
        Role.objects.create(code="admin", name="Administrator")
        Role.objects.create(code="user", name="User")

        # Call the task
        result = tasks.generate_xlsx_from_queryset_task.run(
            app_label="core",
            model_name="Role",
            queryset_filters=None,
            filename="roles_export.xlsx",
            storage_backend="local",
        )

        # Verify result
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["file_url"], "https://example.com/test.xlsx")

        # Verify progress tracker was initialized
        self.assertTrue(mock_tracker_class.called)

        # Verify set_total was called with row count
        mock_tracker.set_total.assert_called_once_with(2)
        mock_tracker.set_completed.assert_called_once()


@override_settings(
    EXPORTER_CELERY_ENABLED=True,
    EXPORTER_STORAGE_BACKEND="local",
    EXPORTER_PROGRESS_CHUNK_SIZE=100,
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    },
)
class GenerateXLSXFromViewsetTaskTests(TestCase):
    """Test cases for generate_xlsx_from_viewset_task."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    def test_task_calls_viewset_get_export_data(self, mock_tracker_class, mock_storage_backend):
        """Test that task calls ViewSet's get_export_data method."""
        # Mock storage
        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        # Mock progress tracker
        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from apps.core.models import Role
        from libs.export_xlsx import tasks

        # Create test data
        Role.objects.create(code="admin", name="Administrator")

        # Call the task with a ViewSet that has custom get_export_data
        viewset_class_path = "libs.tests.test_export_xlsx_mixin.TestExportViewSet"
        request_data = {
            "query_params": {},
            "user_id": None,
        }

        result = tasks.generate_xlsx_from_viewset_task.run(
            viewset_class_path=viewset_class_path,
            request_data=request_data,
            filename="roles_export.xlsx",
            storage_backend="local",
        )

        # Verify result
        self.assertEqual(result["status"], "success")
        self.assertEqual(result["file_url"], "https://example.com/test.xlsx")

        # Verify progress tracker was initialized
        self.assertTrue(mock_tracker_class.called)

        # Verify set_completed was called
        mock_tracker.set_completed.assert_called_once()


@override_settings(
    EXPORTER_CELERY_ENABLED=True,
    EXPORTER_STORAGE_BACKEND="local",
    EXPORTER_PROGRESS_CHUNK_SIZE=100,
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    },
)
class GenerateXLSXTaskTemplateTests(SimpleTestCase):
    """Test cases for Celery tasks with template_name and template_context parameters."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    @patch("libs.export_xlsx.tasks.XLSXGenerator")
    def test_generate_xlsx_task_with_template_name(
        self, mock_generator_class, mock_tracker_class, mock_storage_backend
    ):
        """Test generate_xlsx_task passes template_name to generator."""
        # Mock XLSXGenerator
        mock_generator = Mock()
        mock_file_content = Mock()
        mock_generator.generate.return_value = mock_file_content
        mock_generator_class.return_value = mock_generator

        # Mock storage
        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        # Mock progress tracker
        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from libs.export_xlsx import tasks

        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": "User 1"}],
                }
            ]
        }

        # Call the task with template_name
        result = tasks.generate_xlsx_task.run(
            schema,
            filename="test.xlsx",
            storage_backend="local",
            template_name="my_template.xlsx",
            template_context=None,
        )

        # Verify result
        self.assertEqual(result["status"], "success")

        # Verify generator.generate was called with template_name and template_context
        mock_generator.generate.assert_called_once()
        call_kwargs = mock_generator.generate.call_args[1]
        self.assertEqual(call_kwargs["template_name"], "my_template.xlsx")
        self.assertIsNone(call_kwargs["template_context"])

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    @patch("libs.export_xlsx.tasks.XLSXGenerator")
    def test_generate_xlsx_task_with_template_context(
        self, mock_generator_class, mock_tracker_class, mock_storage_backend
    ):
        """Test generate_xlsx_task passes template_context to generator."""
        mock_generator = Mock()
        mock_file_content = Mock()
        mock_generator.generate.return_value = mock_file_content
        mock_generator_class.return_value = mock_generator

        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from libs.export_xlsx import tasks

        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": "User 1"}],
                }
            ]
        }

        template_context = {"{{ title }}": "Report Title", "{{ date }}": "2025-01-01"}

        # Call the task with template_name and template_context
        result = tasks.generate_xlsx_task.run(
            schema,
            filename="test.xlsx",
            storage_backend="local",
            template_name="my_template.xlsx",
            template_context=template_context,
        )

        self.assertEqual(result["status"], "success")

        # Verify generator.generate was called with template_context
        call_kwargs = mock_generator.generate.call_args[1]
        self.assertEqual(call_kwargs["template_name"], "my_template.xlsx")
        self.assertEqual(call_kwargs["template_context"], template_context)

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    @patch("libs.export_xlsx.tasks.XLSXGenerator")
    def test_generate_xlsx_task_without_template(self, mock_generator_class, mock_tracker_class, mock_storage_backend):
        """Test generate_xlsx_task works without template parameters."""
        mock_generator = Mock()
        mock_file_content = Mock()
        mock_generator.generate.return_value = mock_file_content
        mock_generator_class.return_value = mock_generator

        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from libs.export_xlsx import tasks

        schema = {
            "sheets": [
                {
                    "name": "Test",
                    "headers": ["Name"],
                    "field_names": ["name"],
                    "data": [{"name": "User 1"}],
                }
            ]
        }

        # Call the task without template parameters (default)
        result = tasks.generate_xlsx_task.run(
            schema,
            filename="test.xlsx",
            storage_backend="local",
        )

        self.assertEqual(result["status"], "success")

        # Verify generator.generate was called with None for template params
        call_kwargs = mock_generator.generate.call_args[1]
        self.assertIsNone(call_kwargs["template_name"])
        self.assertIsNone(call_kwargs["template_context"])


@override_settings(
    EXPORTER_CELERY_ENABLED=True,
    EXPORTER_STORAGE_BACKEND="local",
    EXPORTER_PROGRESS_CHUNK_SIZE=100,
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    },
)
class GenerateXLSXFromQuerysetTaskTemplateTests(TestCase):
    """Test cases for generate_xlsx_from_queryset_task with template parameters."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    @patch("libs.export_xlsx.tasks.XLSXGenerator")
    def test_queryset_task_with_template_parameters(
        self, mock_generator_class, mock_tracker_class, mock_storage_backend
    ):
        """Test generate_xlsx_from_queryset_task passes template params to generator."""
        mock_generator = Mock()
        mock_file_content = Mock()
        mock_generator.generate.return_value = mock_file_content
        mock_generator_class.return_value = mock_generator

        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from apps.core.models import Role
        from libs.export_xlsx import tasks

        # Create test data
        Role.objects.create(code="admin", name="Administrator")

        template_context = {"{{ company }}": "Test Company"}

        # Call the task with template parameters
        result = tasks.generate_xlsx_from_queryset_task.run(
            app_label="core",
            model_name="Role",
            queryset_filters=None,
            filename="roles_export.xlsx",
            storage_backend="local",
            template_name="queryset_template.xlsx",
            template_context=template_context,
        )

        self.assertEqual(result["status"], "success")

        # Verify generator.generate was called with template params
        call_kwargs = mock_generator.generate.call_args[1]
        self.assertEqual(call_kwargs["template_name"], "queryset_template.xlsx")
        self.assertEqual(call_kwargs["template_context"], template_context)


@override_settings(
    EXPORTER_CELERY_ENABLED=True,
    EXPORTER_STORAGE_BACKEND="local",
    EXPORTER_PROGRESS_CHUNK_SIZE=100,
    CACHES={
        "default": {
            "BACKEND": "django.core.cache.backends.locmem.LocMemCache",
            "LOCATION": "test-progress-cache",
        }
    },
)
class GenerateXLSXFromViewsetTaskTemplateTests(TestCase):
    """Test cases for generate_xlsx_from_viewset_task with template parameters."""

    def setUp(self):
        """Set up test fixtures."""
        cache.clear()

    def tearDown(self):
        """Clean up after tests."""
        cache.clear()

    @patch("libs.export_xlsx.tasks.get_storage_backend")
    @patch("libs.export_xlsx.tasks.ExportProgressTracker")
    @patch("libs.export_xlsx.tasks.XLSXGenerator")
    def test_viewset_task_with_template_parameters(
        self, mock_generator_class, mock_tracker_class, mock_storage_backend
    ):
        """Test generate_xlsx_from_viewset_task passes template params to generator."""
        mock_generator = Mock()
        mock_file_content = Mock()
        mock_generator.generate.return_value = mock_file_content
        mock_generator_class.return_value = mock_generator

        mock_storage = Mock()
        mock_storage.save.return_value = "exports/test.xlsx"
        mock_storage.get_url.return_value = "https://example.com/test.xlsx"
        mock_storage_backend.return_value = mock_storage

        mock_tracker = Mock()
        mock_tracker_class.return_value = mock_tracker

        from apps.core.models import Role
        from libs.export_xlsx import tasks

        Role.objects.create(code="admin", name="Administrator")

        viewset_class_path = "libs.tests.test_export_xlsx_mixin.TestExportViewSet"
        request_data = {
            "query_params": {},
            "user_id": None,
        }
        template_context = {"{{ report_name }}": "ViewSet Export Report"}

        result = tasks.generate_xlsx_from_viewset_task.run(
            viewset_class_path=viewset_class_path,
            request_data=request_data,
            filename="roles_export.xlsx",
            storage_backend="local",
            template_name="viewset_template.xlsx",
            template_context=template_context,
        )

        self.assertEqual(result["status"], "success")

        # Verify generator.generate was called with template params
        call_kwargs = mock_generator.generate.call_args[1]
        self.assertEqual(call_kwargs["template_name"], "viewset_template.xlsx")
        self.assertEqual(call_kwargs["template_context"], template_context)
