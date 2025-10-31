"""
Error report generator for XLSX import.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ErrorReportGenerator:
    """
    Generate XLSX error reports for import failures.
    """

    def generate(  # noqa: C901
        self, errors: list[dict], original_data: list[list] | None = None, headers: list[str] | None = None
    ) -> bytes:
        """
        Generate error report as XLSX file.

        Args:
            errors: List of error dictionaries with 'row' and 'errors' keys
            original_data: Optional original data rows for reference
            headers: Optional header names

        Returns:
            bytes: XLSX file content
        """
        workbook = Workbook()

        # Error Summary Sheet
        summary_sheet = workbook.active
        summary_sheet.title = "Error Summary"

        # Header styling
        header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Add headers
        summary_sheet["A1"] = "Row Number"
        summary_sheet["B1"] = "Field"
        summary_sheet["C1"] = "Error Message"

        # Style headers
        for cell in ["A1", "B1", "C1"]:
            summary_sheet[cell].fill = header_fill
            summary_sheet[cell].font = header_font

        # Add error data
        row_num = 2
        for error_item in errors:
            row = error_item.get("row", "Unknown")
            errors_dict = error_item.get("errors", {})

            if isinstance(errors_dict, dict):
                for field, messages in errors_dict.items():
                    if isinstance(messages, list):
                        for message in messages:
                            summary_sheet[f"A{row_num}"] = row
                            summary_sheet[f"B{row_num}"] = field
                            summary_sheet[f"C{row_num}"] = str(message)
                            row_num += 1
                    else:
                        summary_sheet[f"A{row_num}"] = row
                        summary_sheet[f"B{row_num}"] = field
                        summary_sheet[f"C{row_num}"] = str(messages)
                        row_num += 1

        # Auto-adjust column widths
        for column in ["A", "B", "C"]:
            max_length = 0
            for cell in summary_sheet[column]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            summary_sheet.column_dimensions[column].width = min(max_length + 2, 50)

        # If original data provided, add it to a separate sheet
        if original_data and headers:
            data_sheet = workbook.create_sheet(title="Original Data with Errors")

            # Add headers
            for col_idx, header in enumerate(headers, start=1):
                cell = data_sheet.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Add "Error" column
            error_col = len(headers) + 1
            error_cell = data_sheet.cell(row=1, column=error_col, value="Error")
            error_cell.fill = header_fill
            error_cell.font = header_font

            # Add data rows with error indicators
            error_rows = {e["row"] for e in errors}

            for row_idx, data_row in enumerate(original_data, start=2):
                # Add data columns
                for col_idx, value in enumerate(data_row, start=1):
                    data_sheet.cell(row=row_idx, column=col_idx, value=value)

                # Mark if row has error
                if row_idx in error_rows:
                    error_messages = []
                    for error_item in errors:
                        if error_item.get("row") == row_idx:
                            errors_dict = error_item.get("errors", {})
                            for field, messages in errors_dict.items():
                                if isinstance(messages, list):
                                    error_messages.extend([f"{field}: {m}" for m in messages])
                                else:
                                    error_messages.append(f"{field}: {messages}")

                    error_text = "; ".join(error_messages)
                    cell = data_sheet.cell(row=row_idx, column=error_col, value=error_text)
                    # Highlight error row
                    for col_idx in range(1, error_col + 1):
                        data_sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
                        )

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
