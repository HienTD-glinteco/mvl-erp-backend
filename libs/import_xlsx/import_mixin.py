"""
Mixin for adding XLSX import functionality to DRF ViewSets.

This module provides a reusable ImportXLSXMixin that adds a universal import action
to any DRF ViewSet, enabling data import from Excel files (.xlsx) into Django models.
"""

import logging
from typing import Any

from django.conf import settings
from django.db.models import QuerySet
from django.utils.translation import gettext as _
from drf_spectacular.utils import OpenApiParameter, OpenApiResponse, extend_schema
from openpyxl import load_workbook
from rest_framework import exceptions, status
from rest_framework.decorators import action
from rest_framework.request import Request
from rest_framework.response import Response

from .error_report import ErrorReportGenerator
from .import_constants import (
    ERROR_ASYNC_NOT_ENABLED,
    ERROR_EMPTY_FILE,
    ERROR_INVALID_FILE_TYPE,
    ERROR_NO_FILE,
    SUCCESS_IMPORT_COMPLETE,
    SUCCESS_PREVIEW_COMPLETE,
)
from .mapping_config import MappingConfigParser
from .multi_model_processor import MultiModelProcessor
from .serializers import ImportAsyncResponseSerializer, ImportResultSerializer
from .storage import get_storage_backend
from .tasks import import_xlsx_task

logger = logging.getLogger(__name__)


class ImportXLSXMixin:
    """
    Mixin for DRF ViewSets to add XLSX import functionality.

    This mixin provides a universal `/import/` action that accepts XLSX files,
    maps columns to model fields, validates data, and performs bulk create/update.

    By default, validation is done at the model level (using model.full_clean()).
    This allows importing all model fields including read-only or auto-generated ones.

    If you need serializer-level validation, override get_import_serializer_class()
    to return a custom import serializer without read-only field restrictions.

    Example 1 - Basic import with model-level validation:
        class ProjectViewSet(ImportXLSXMixin, ModelViewSet):
            queryset = Project.objects.all()
            serializer_class = ProjectSerializer

    Example 2 - Custom import schema:
        class ProjectViewSet(ImportXLSXMixin, ModelViewSet):
            queryset = Project.objects.all()
            serializer_class = ProjectSerializer

            def get_import_schema(self, request, file):
                return {
                    "fields": ["name", "start_date", "budget"],
                    "required": ["name"],
                }

    Example 3 - Custom import serializer for validation:
        class ProjectImportSerializer(serializers.ModelSerializer):
            class Meta:
                model = Project
                fields = ["name", "code", "budget"]
                # No read-only fields, all can be imported

        class ProjectViewSet(ImportXLSXMixin, ModelViewSet):
            queryset = Project.objects.all()
            serializer_class = ProjectSerializer

            def get_import_serializer_class(self):
                return ProjectImportSerializer
    """

    queryset: QuerySet[Any]

    @extend_schema(
        summary="Import data from XLSX",
        description="Import model instances from an Excel (.xlsx) file. "
        "Supports async processing, preview mode, and error report download.",
        request={
            "multipart/form-data": {
                "type": "object",
                "properties": {
                    "file": {
                        "type": "string",
                        "format": "binary",
                        "description": "XLSX file to import",
                    },
                },
                "required": ["file"],
            }
        },
        parameters=[
            OpenApiParameter(
                name="async",
                description="If 'true', process import in background using Celery (requires IMPORTER_CELERY_ENABLED=true)",
                required=False,
                type=bool,
            ),
            OpenApiParameter(
                name="preview",
                description="If 'true', validate data without saving (dry-run mode)",
                required=False,
                type=bool,
            ),
        ],
        responses={
            200: ImportResultSerializer,
            202: ImportAsyncResponseSerializer,
            400: OpenApiResponse(description="Bad request"),
        },
    )
    @action(detail=False, methods=["post"])
    def import_data(self, request: Request) -> Response:
        """
        Import data from XLSX file.

        Query parameters:
            async: If 'true', process in background (requires Celery)
            preview: If 'true', validate without saving (dry-run)

        This action:
        1. Validates the uploaded file
        2. Loads the import schema (auto or custom)
        3. Parses and validates each row
        4. (If not preview) Bulk creates/updates model instances
        5. Returns import summary with errors
        6. (If errors) Optionally generates error report XLSX

        Returns:
            - Synchronous (200): Import results with error details
            - Asynchronous (202): Task ID and status information
            - Preview (200): Validation results without saving
        """
        try:
            file = self._validate_file(request)
            use_async, preview_mode, celery_enabled = self._get_import_modes(request)

            if use_async and not celery_enabled:
                return Response(
                    {"error": _(ERROR_ASYNC_NOT_ENABLED)},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        except (ValueError, exceptions.UnsupportedMediaType) as e:
            # Handle validation errors (file not found, invalid type, unsupported media, etc.)
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            logger.exception(f"Import failed: {e}")
            return Response(
                {"detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            # Check if using advanced config-driven import
            import_config = self.get_import_config(request, file)

            if import_config:
                # Use advanced multi-model processor
                return self._handle_advanced_import(
                    request=request,
                    file=file,
                    config=import_config,
                    use_async=use_async,
                    preview_mode=preview_mode,
                )

            # Otherwise, use simple schema-based import
            # Load import schema
            import_schema = self.get_import_schema(request, file)

            if use_async and not preview_mode:
                # Handle async import
                return self._handle_async_import(request, file, import_schema)

            # Parse and validate data
            parsed_data, errors, original_data, headers = self._parse_xlsx_file_with_original(file, import_schema)

            if not parsed_data and not errors:
                return Response(
                    {"detail": _(ERROR_EMPTY_FILE)},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            if preview_mode:
                # Preview mode - don't save, just return validation results
                return self._handle_preview(parsed_data, errors)

            # Import data
            success_count = 0
            if parsed_data:
                success_count = self._bulk_import(parsed_data, import_schema, errors)

            # Generate error report if errors exist
            error_file_url = None
            if errors:
                error_file_url = self._generate_error_report(errors, original_data, headers)

            response_data = {
                "success_count": success_count,
                "error_count": len(errors),
                "errors": errors[:100],  # Limit errors in response
                "detail": _(SUCCESS_IMPORT_COMPLETE),
            }

            if error_file_url:
                response_data["error_file_url"] = error_file_url

            return Response(response_data, status=status.HTTP_200_OK)

        except ValueError as e:
            logger.error(f"Import validation error: {e}")
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            logger.exception(f"Import failed: {e}")
            return Response(
                {"detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def _validate_file(self, request: Request):
        # Validate file upload
        if "file" not in request.FILES:
            raise ValueError(_(ERROR_NO_FILE))

        file = request.FILES["file"]

        if not file.name.endswith(".xlsx"):
            raise ValueError(_(ERROR_INVALID_FILE_TYPE))

        return file

    def _get_import_modes(self, request: Request):
        # Check for async mode
        use_async = request.query_params.get("async", "false").lower() == "true"
        celery_enabled = getattr(settings, "IMPORTER_CELERY_ENABLED", False)

        # Check for preview mode
        preview_mode = request.query_params.get("preview", "false").lower() == "true"

        return use_async, preview_mode, celery_enabled

    def get_import_schema(self, request: Request, file: Any) -> dict:
        """
        Get the import schema for mapping columns to fields.

        Override this method to customize the import schema.
        By default, auto-generates schema from model fields.

        Args:
            request: The HTTP request
            file: The uploaded file

        Returns:
            dict: Import schema with keys:
                - fields: List of field names to import
                - required: List of required field names
                - validators: Optional dict of field validators
        """
        return self._auto_generate_schema()

    def get_import_serializer_class(self):
        """
        Get the serializer class to use for import validation.

        Override this method to use a different serializer for imports
        than for regular CRUD operations. This is useful when the regular
        serializer has read-only fields (like auto-generated codes) or
        required fields (like relations) that should not be required for imports.

        Returns:
            Serializer class or None to skip serializer validation
        """
        return None  # By default, use model-level validation only

    def get_import_config(self, request: Request, file: Any) -> dict | None:
        """
        Get advanced import configuration for multi-model imports.

        Override this method to provide a full mapping configuration
        that supports multiple models, field combinations, nested relationships,
        and conditional relations.

        If this method returns a configuration, it will be used instead of
        get_import_schema(), enabling advanced import features.

        Returns:
            dict or None: Configuration dictionary (see MappingConfigParser for format)
                or None to use simple schema-based import

        Example:
            def get_import_config(self, request, file):
                return {
                    "sheets": [{
                        "name": "Employees",
                        "model": "Employee",
                        "app_label": "hrm",
                        "fields": {
                            "employee_code": "Employee Code",
                            "start_date": {
                                "combine": ["Day", "Month", "Year"],
                                "format": "YYYY-MM-DD"
                            },
                            "department": {
                                "model": "Department",
                                "lookup": "Department",
                                "create_if_not_found": true,
                                "relations": {
                                    "division": {
                                        "model": "Division",
                                        "lookup": "Division"
                                    }
                                }
                            }
                        },
                        "relations": {
                            "accounts": [{
                                "model": "Account",
                                "fields": {
                                    "bank": "VPBank",
                                    "account_number": "VPBank Account"
                                }
                            }]
                        }
                    }]
                }
        """
        return None  # By default, use simple schema-based import

    def _handle_async_import(self, request: Request, file: Any, schema: dict) -> Response:
        """
        Handle async import using Celery.

        Args:
            request: HTTP request
            file: Uploaded file
            schema: Import schema

        Returns:
            Response: Async task information
        """
        # Save file temporarily
        storage = get_storage_backend()
        file_content = file.read()
        file_path = storage.save_temp_file(file_content, file.name)

        # Get model info
        model = self.queryset.model
        app_label = model._meta.app_label
        model_name = model.__name__

        # Get user ID for audit logging
        user_id = request.user.id if hasattr(request, "user") and request.user.is_authenticated else None

        # Trigger background task
        task = import_xlsx_task.delay(app_label, model_name, file_path, schema, user_id)

        return Response(
            {
                "task_id": task.id,
                "status": "PENDING",
                "message": _("Import task has been queued for processing"),
            },
            status=status.HTTP_202_ACCEPTED,
        )

    def _handle_preview(self, parsed_data: list[dict], errors: list[dict]) -> Response:
        """
        Handle preview mode (dry-run).

        Args:
            parsed_data: Validated data
            errors: Validation errors

        Returns:
            Response: Preview results
        """
        max_preview = getattr(settings, "IMPORTER_MAX_PREVIEW_ROWS", 10)
        preview_data = parsed_data[:max_preview] if parsed_data else []

        return Response(
            {
                "valid_count": len(parsed_data),
                "invalid_count": len(errors),
                "errors": errors[:100],  # Limit errors in response
                "preview_data": preview_data[:10],  # Show first 10 valid rows
                "detail": _(SUCCESS_PREVIEW_COMPLETE),
            },
            status=status.HTTP_200_OK,
        )

    def _generate_error_report(self, errors: list[dict], original_data: list[list], headers: list[str]) -> str:
        """
        Generate error report XLSX file.

        Args:
            errors: List of errors
            original_data: Original data rows
            headers: Column headers

        Returns:
            str: URL to download error report
        """
        try:
            generator = ErrorReportGenerator()
            report_content = generator.generate(errors, original_data, headers)

            # Save to storage
            storage = get_storage_backend()
            file_path = storage.save_error_report(report_content)

            # Get URL
            return storage.get_url(file_path)

        except Exception as e:
            logger.warning(f"Failed to generate error report: {e}")
            return None  # type: ignore[return-value]

    def _auto_generate_schema(self) -> dict:
        """
        Auto-generate import schema from model fields.

        Ignores fields like id, created_at, updated_at, and AutoFields.

        Returns:
            dict: Auto-generated import schema
        """
        from django.db import models

        from .import_constants import IGNORED_FIELD_NAMES

        model = self.queryset.model
        fields = []
        required = []

        for field in model._meta.get_fields():
            # Skip ignored fields
            if field.name in IGNORED_FIELD_NAMES:
                continue

            # Skip auto fields
            if isinstance(field, (models.AutoField, models.BigAutoField)):
                continue

            # Skip reverse relations
            if field.many_to_many or field.one_to_many or field.one_to_one:
                continue

            # Add field to schema
            fields.append(field.name)

            # Check if field is required
            if hasattr(field, "blank") and not field.blank and not field.null:
                required.append(field.name)

        return {
            "fields": fields,
            "required": required,
        }

    def _parse_xlsx_file(self, file: Any, schema: dict) -> tuple[list[dict], list[dict]]:
        """
        Parse XLSX file and validate data.

        Args:
            file: The uploaded XLSX file
            schema: Import schema

        Returns:
            tuple: (parsed_data, errors)
                - parsed_data: List of validated dictionaries
                - errors: List of error dictionaries with row numbers
        """
        parsed_data, errors, _, _ = self._parse_xlsx_file_with_original(file, schema)
        return parsed_data, errors

    def _parse_xlsx_file_with_original(
        self, file: Any, schema: dict
    ) -> tuple[list[dict], list[dict], list[list], list[str]]:
        """
        Parse XLSX file and validate data, keeping original data for error reports.

        Args:
            file: The uploaded XLSX file
            schema: Import schema

        Returns:
            tuple: (parsed_data, errors, original_data, headers)
                - parsed_data: List of validated dictionaries
                - errors: List of error dictionaries with row numbers
                - original_data: Original data rows (for error reporting)
                - headers: Column headers
        """
        parsed_data = []
        errors = []
        original_data = []
        headers = []

        try:
            workbook = load_workbook(file, read_only=True)
            sheet = workbook.active

            # Get header row
            headers = self._extract_headers(sheet)

            # Map headers to fields
            field_mapping = self._map_headers_to_fields(headers, schema["fields"])

            # Parse data rows
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue

                # Store original row data
                original_data.append(list(row))

                # Process single row
                row_data, row_errors = self._process_row(row, headers, field_mapping, schema)

                if row_errors:
                    errors.append({"row": row_idx, "errors": row_errors})
                elif row_data:
                    parsed_data.append(row_data)

            workbook.close()

        except Exception as e:
            logger.exception(f"Error parsing XLSX file: {e}")
            errors.append({"row": 0, "errors": {"_general": str(e)}})

        return parsed_data, errors, original_data, headers

    def _extract_headers(self, sheet: Any) -> list[str]:
        """
        Extract header row from worksheet.

        Args:
            sheet: Worksheet object

        Returns:
            list: List of header names
        """
        from .utils import extract_headers

        return extract_headers(sheet)

    def _process_row(
        self, row: tuple, headers: list[str], field_mapping: dict, schema: dict
    ) -> tuple[dict | None, dict]:
        """
        Process a single row of data.

        Args:
            row: Tuple of cell values
            headers: List of header names
            field_mapping: Mapping of headers to field names
            schema: Import schema

        Returns:
            tuple: (validated_data, errors)
        """
        from .utils import process_row

        model = self.queryset.model
        row_data, row_errors = process_row(row, headers, field_mapping, schema, model)

        # Validate data using import serializer if provided
        if not row_errors:
            serializer_class = self.get_import_serializer_class()

            if serializer_class:
                # Use import-specific serializer for validation
                try:
                    serializer = serializer_class(data=row_data)
                    if serializer.is_valid():
                        return serializer.validated_data, {}
                    else:
                        row_errors.update(serializer.errors)
                except Exception as e:
                    row_errors["_general"] = str(e)
            else:
                # No import serializer - use model-level validation only
                # Just return the raw data, validation will happen at model.full_clean()
                return row_data, {}

        return None, row_errors

    def _map_headers_to_fields(self, headers: list[str], fields: list[str]) -> dict:
        """
        Map Excel headers to model field names.

        Performs case-insensitive matching and handles spaces/underscores.

        Args:
            headers: List of column headers from Excel
            fields: List of valid field names from schema

        Returns:
            dict: Mapping of header to field name
        """
        from .utils import map_headers_to_fields

        return map_headers_to_fields(headers, fields)

    def _bulk_import(self, data: list[dict], schema: dict, errors: list[dict]) -> int:
        """
        Bulk import validated data into database.

        Args:
            data: List of validated data dictionaries
            schema: Import schema
            errors: List to append errors to

        Returns:
            int: Number of successfully imported records
        """
        from .utils import bulk_import_data

        model = self.queryset.model
        request = self.request if hasattr(self, "request") else None
        user_id = request.user.id if request and hasattr(request, "user") and request.user.is_authenticated else None

        # Use shared utility with request context for audit logging
        return bulk_import_data(model, data, errors, user_id=user_id, request=request)

    def _handle_advanced_import(
        self,
        request: Request,
        file: Any,
        config: dict,
        use_async: bool = False,
        preview_mode: bool = False,
    ) -> Response:
        """
        Handle advanced config-driven import with multi-model support.

        Args:
            request: HTTP request
            file: Uploaded XLSX file
            config: Import configuration dictionary
            use_async: Whether to process asynchronously
            preview_mode: Whether to run in preview mode

        Returns:
            Response: Import results or async task info
        """
        try:
            # Validate and parse configuration
            config_parser = MappingConfigParser(config)

            # TODO: Implement async handling for advanced imports
            if use_async:
                return Response(
                    {"detail": "Async mode not yet supported for advanced imports"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Load workbook
            workbook = load_workbook(file, data_only=True)

            # Process using MultiModelProcessor
            processor = MultiModelProcessor(config_parser)
            results = processor.process_file(
                workbook=workbook,
                preview=preview_mode,
                user=request.user if hasattr(request, "user") else None,
                request=request,
            )

            # Build response
            response_data = {
                "success_count": results["success_count"],
                "error_count": results["error_count"],
                "errors": results["errors"],
                "detail": _(SUCCESS_PREVIEW_COMPLETE if preview_mode else SUCCESS_IMPORT_COMPLETE),
            }

            # Add preview data if in preview mode
            if preview_mode and results.get("preview_data"):
                response_data["preview_data"] = results["preview_data"][:10]  # Limit to 10 rows
                response_data["valid_count"] = results["success_count"]
                response_data["invalid_count"] = results["error_count"]

            # Generate error report if errors exist and not in preview mode
            if not preview_mode and results["errors"]:
                try:
                    error_generator = ErrorReportGenerator()
                    error_file_path = error_generator.generate_report(  # type: ignore[attr-defined]
                        errors=results["errors"],
                        original_data=[],  # TODO: Get original data from processor
                        headers=[],
                    )

                    # Get storage backend
                    storage = get_storage_backend()
                    error_file_url = storage.get_url(error_file_path)
                    response_data["error_file_url"] = error_file_url
                except Exception as e:
                    logger.warning(f"Failed to generate error report: {e}")

            return Response(response_data, status=status.HTTP_200_OK)

        except ValueError as e:
            logger.error(f"Configuration error: {e}")
            return Response(
                {"detail": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception as e:
            logger.error(f"Import error: {e}")
            return Response(
                {"detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
