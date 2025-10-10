"""
Celery tasks for async XLSX import.
"""

import logging

from celery import shared_task
from django.apps import apps
from django.utils.translation import gettext as _
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@shared_task(bind=True, name="import_xlsx.import_file")
def import_xlsx_task(self, app_label, model_name, file_path, schema, user_id=None):
    """
    Background task to import XLSX file.

    Args:
        app_label: Django app label (e.g., 'core')
        model_name: Model name (e.g., 'Role')
        file_path: Path to uploaded XLSX file
        schema: Import schema definition
        user_id: Optional user ID for audit logging

    Returns:
        dict: Result with keys:
            - status: 'success' or 'error'
            - success_count: Number of successfully imported rows
            - error_count: Number of errors
            - errors: List of error details
            - error_file_url: URL to download error report (if errors exist)
    """
    from django.core.files.storage import default_storage
    from django.db import transaction

    from .error_report import ErrorReportGenerator
    from .storage import get_storage_backend

    try:
        # Get model class
        model = apps.get_model(app_label, model_name)

        # Load the file from storage
        if not default_storage.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        # Open and parse XLSX file
        with default_storage.open(file_path, "rb") as f:
            workbook = load_workbook(f, read_only=True)
            sheet = workbook.active

            # Extract headers
            headers = []
            for cell in sheet[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())

            # Map headers to fields
            field_mapping = _map_headers_to_fields(headers, schema["fields"])

            # Parse and validate data
            parsed_data = []
            errors = []
            original_data = []

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                # Skip empty rows
                if not any(row):
                    continue

                # Store original row data
                original_data.append(list(row))

                # Process row
                row_data, row_errors = _process_row(row, headers, field_mapping, schema, model)

                if row_errors:
                    errors.append({"row": row_idx, "errors": row_errors})
                elif row_data:
                    parsed_data.append(row_data)

            workbook.close()

        # Import data
        success_count = 0
        if parsed_data:
            success_count = _bulk_import_data(model, parsed_data, errors, user_id)

        # Generate error report if errors exist
        error_file_url = None
        if errors:
            try:
                storage = get_storage_backend()
                generator = ErrorReportGenerator()
                report_content = generator.generate(errors, original_data, headers)
                report_path = storage.save_error_report(report_content)
                error_file_url = storage.get_url(report_path)
            except Exception as e:
                logger.warning(f"Failed to generate error report: {e}")

        # Clean up temporary file
        try:
            default_storage.delete(file_path)
        except Exception as e:
            logger.warning(f"Failed to delete temp file: {e}")

        return {
            "status": "success",
            "success_count": success_count,
            "error_count": len(errors),
            "errors": errors[:100],  # Limit errors in response
            "error_file_url": error_file_url,
        }

    except Exception as e:
        logger.exception(f"Import task failed: {e}")
        return {
            "status": "error",
            "error": str(e),
            "success_count": 0,
            "error_count": 0,
            "errors": [],
        }


def _map_headers_to_fields(headers: list[str], fields: list[str]) -> dict:
    """
    Map Excel headers to model field names.

    Args:
        headers: List of column headers from Excel
        fields: List of valid field names from schema

    Returns:
        dict: Mapping of header to field name
    """
    mapping = {}
    field_lookup = {field.lower().replace("_", " "): field for field in fields}

    for header in headers:
        normalized_header = header.lower().replace("_", " ")
        if normalized_header in field_lookup:
            mapping[header] = field_lookup[normalized_header]
        elif header in fields:
            mapping[header] = header

    return mapping


def _process_row(row: tuple, headers: list[str], field_mapping: dict, schema: dict, model) -> tuple[dict | None, dict]:
    """
    Process a single row of data.

    Args:
        row: Tuple of cell values
        headers: List of header names
        field_mapping: Mapping of headers to field names
        schema: Import schema
        model: Django model class

    Returns:
        tuple: (validated_data, errors)
    """
    from django.db import models as django_models

    row_data = {}
    row_errors = {}
    m2m_data = {}

    # Map cells to fields
    for col_idx, cell_value in enumerate(row):
        if col_idx < len(headers):
            field_name = field_mapping.get(headers[col_idx])
            if field_name:
                try:
                    model_field = model._meta.get_field(field_name)

                    if isinstance(model_field, django_models.ForeignKey):
                        # Resolve ForeignKey
                        try:
                            row_data[field_name] = _resolve_foreign_key(model_field, cell_value)
                        except ValueError as e:
                            row_errors[field_name] = str(e)
                    elif isinstance(model_field, django_models.ManyToManyField):
                        # Store ManyToMany for later
                        m2m_data[field_name] = _resolve_many_to_many(model_field, cell_value)
                    else:
                        # Regular field
                        row_data[field_name] = cell_value
                except:
                    # Field not found in model, just store as-is
                    row_data[field_name] = cell_value

    # Store M2M data for later use
    if m2m_data:
        row_data["_m2m_data"] = m2m_data

    # Validate required fields
    for required_field in schema.get("required", []):
        if not row_data.get(required_field):
            row_errors[required_field] = _("This field is required")

    if row_errors:
        return None, row_errors

    return row_data, {}


def _resolve_foreign_key(field, value):
    """Resolve ForeignKey value."""
    if value is None or value == "":
        return None

    related_model = field.related_model

    # Try to find by primary key first
    try:
        if isinstance(value, (int, float)):
            return related_model.objects.get(pk=int(value))
    except (related_model.DoesNotExist, ValueError):
        pass

    # Try to find by common natural keys
    for lookup_field in ["name", "code", "email", "username"]:
        if hasattr(related_model, lookup_field):
            try:
                return related_model.objects.get(**{lookup_field: str(value)})
            except related_model.DoesNotExist:
                continue

    raise ValueError(f"Related object not found for {field.name}: {value}")


def _resolve_many_to_many(field, value):
    """Resolve ManyToMany values."""
    if value is None or value == "":
        return []

    related_model = field.related_model
    values = [v.strip() for v in str(value).split(",") if v.strip()]
    instances = []

    for val in values:
        try:
            # Try by primary key
            if val.isdigit():
                instances.append(related_model.objects.get(pk=int(val)))
                continue

            # Try natural keys
            for lookup_field in ["name", "code", "email", "username"]:
                if hasattr(related_model, lookup_field):
                    try:
                        instances.append(related_model.objects.get(**{lookup_field: val}))
                        break
                    except related_model.DoesNotExist:
                        continue
        except Exception as e:
            logger.warning(f"Could not resolve value: {val} - {e}")

    return instances


def _bulk_import_data(model, data: list[dict], errors: list[dict], user_id=None) -> int:
    """
    Bulk import validated data into database.

    Args:
        model: Django model class
        data: List of validated data dictionaries
        errors: List to append errors to
        user_id: Optional user ID for audit logging

    Returns:
        int: Number of successfully imported records
    """
    from django.db import transaction

    success_count = 0

    try:
        with transaction.atomic():
            for item_data in data:
                try:
                    # Extract M2M data if present
                    m2m_data = item_data.pop("_m2m_data", {})

                    # Create model instance
                    instance = model(**item_data)
                    instance.full_clean()
                    instance.save()

                    # Set ManyToMany relationships
                    for field_name, related_objects in m2m_data.items():
                        getattr(instance, field_name).set(related_objects)

                    success_count += 1

                    # Log audit event if audit logging is enabled
                    _log_import_audit(instance, user_id)

                except Exception as e:
                    logger.warning(f"Failed to import row: {e}")
                    errors.append(
                        {
                            "row": success_count + len(errors) + 2,
                            "errors": {"_general": str(e)},
                        }
                    )

    except Exception as e:
        logger.exception(f"Bulk import failed: {e}")
        raise

    return success_count


def _log_import_audit(instance, user_id=None):
    """
    Log audit event for imported instance.

    Args:
        instance: The imported model instance
        user_id: Optional user ID
    """
    try:
        from apps.audit_logging import LogAction, log_audit_event
        from apps.core.models import User

        user = None
        if user_id:
            try:
                user = User.objects.get(pk=user_id)
            except User.DoesNotExist:
                pass

        log_audit_event(
            action=LogAction.IMPORT,
            modified_object=instance,
            user=user,
            request=None,
        )
    except ImportError:
        # Audit logging not available
        pass
    except Exception as e:
        logger.warning(f"Failed to log audit event: {e}")
