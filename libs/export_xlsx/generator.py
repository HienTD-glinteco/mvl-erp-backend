"""
XLSX generator for creating Excel files from schema definitions.
"""

import logging
import re
import time
from io import BytesIO

from django.conf import settings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from openpyxl.worksheet.worksheet import Worksheet

from .constants import (
    DEFAULT_HEADER_ALIGNMENT,
    DEFAULT_HEADER_BG_COLOR,
    DEFAULT_HEADER_FONT_BOLD,
    DEFAULT_HEADER_FONT_SIZE,
    ERROR_INVALID_SCHEMA,
)

logger = logging.getLogger(__name__)


class XLSXGenerator:
    """
    Generator for creating XLSX files from schema definitions.
    """

    def __init__(self, progress_callback=None, chunk_size=500):
        """
        Initialize XLSX generator.

        Args:
            progress_callback: Optional callback function(rows_processed: int) for progress updates
            chunk_size: Number of rows to process before calling progress_callback
        """
        self.workbook = None
        self.progress_callback = progress_callback
        self.chunk_size = chunk_size
        self.total_rows_processed = 0

    def generate(self, schema, template_name=None, template_context=None):
        """
        Generate XLSX file from schema.

        Args:
            schema: Export schema with structure:
                {
                    "sheets": [{
                        "name": str,
                        "headers": [str, ...],
                        "field_names": [str, ...],  # Optional
                        "data": [dict, ...],
                        "groups": [{"title": str, "span": int}, ...],  # Optional
                        "merge_rules": [str, ...]  # Optional
                    }]
                }

        Returns:
            BytesIO: Excel file content
        """
        if not schema or "sheets" not in schema:
            raise ValueError(ERROR_INVALID_SCHEMA)

        # NOTE: Only use the template_name if it's a valid xlsx or xls file, else just normal generation.
        if not template_name or (not template_name.endswith(".xlsx") and not template_name.endswith(".xls")):
            self.workbook = Workbook()
            # Remove default sheet
            if "Sheet" in self.workbook.sheetnames:
                del self.workbook["Sheet"]

            # Calculate total rows for progress tracking
            total_rows = self._calculate_total_rows(schema)

            # Create each sheet
            for sheet_def in schema["sheets"]:
                self._create_sheet(sheet_def)
        else:
            self.workbook = self._generate_workbook_from_template(schema, template_name, template_context)

        # Save to BytesIO
        output = BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output

    def _generate_workbook_from_template(self, schema, template_name, template_context=None):
        """
        Generate workbook from template file.

        Args:
            schema: Export schema
            template_name: Template file name
            context: Context for template rendering, could be None if not needed for rendering template
        Returns:
            Workbook: Generated workbook
        """
        workbook = load_workbook(filename=template_name, data_only=True)
        sheet: Worksheet = self._render_sheet_with_context(workbook.active, template_context)

        # NOTE: use template means only one sheet supported for now
        sheet_def = schema["sheets"][0]
        start_row = sheet.max_row + 1
        current_row = start_row if start_row > 9 else 9  # NOTE: this usually the line right under the logo image
        self._finalize_sheet(sheet, sheet_def, current_row)
        self._adjust_sheet_title(sheet)
        return workbook

    def _render_sheet_with_context(self, sheet: Worksheet, template_context: dict = None) -> Worksheet:  # type: ignore
        """
        Render sheet by replacing placeholders with context values.

        Args:
            sheet: Worksheet object
            context: Context dictionary for rendering
        """
        if not template_context:
            return sheet

        # NOTE: support only simple placeholders like {{ key }} or {{key}} - just like Django templates
        supported_placeholder_pre_post_fixes = [
            ("{{ ", " }}"),
            ("{{", "}}"),
        ]
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                try:
                    # NOTE: replace cell with correct formatted placeholder with context value.
                    valid = False
                    for pre, post in supported_placeholder_pre_post_fixes:
                        if isinstance(cell.value, str) and cell.value.startswith(pre) and cell.value.endswith(post):
                            valid = True
                            break

                    if not valid or cell.value not in template_context:
                        # NOTE: invalid placeholder will be skipped, avoid breaking flow.
                        continue

                    cell.value = template_context[cell.value]
                except Exception as e:
                    # NOTE: Any error during rendering a cell will be skipped.
                    logger.debug(f"Skipping cell due to error: {e}")
        return sheet

    def _calculate_total_rows(self, schema):
        """
        Calculate total number of data rows in schema.

        Args:
            schema: Export schema

        Returns:
            int: Total number of data rows
        """
        total = 0
        for sheet_def in schema.get("sheets", []):
            data = sheet_def.get("data", [])
            total += len(data)
        return total

    def _create_sheet(self, sheet_def):
        """
        Create a single sheet from definition.

        Args:
            sheet_def: Sheet definition dictionary
        """
        # Create worksheet
        ws = self.workbook.create_sheet()

        self._finalize_sheet(ws, sheet_def)

    def _finalize_sheet(
        self,
        ws: Worksheet,
        sheet_def: dict,
        current_row: int = 1,
    ):
        """
        Finalize worksheet by adding headers, data, and applying styles.

        This is mutual function for both template-based and schema-based sheet creation.
        """

        title = re.sub(INVALID_TITLE_REGEX, "-", sheet_def.get("name", "Sheet1"))
        headers = sheet_def.get("headers", [])
        field_names = sheet_def.get("field_names", headers)
        data = sheet_def.get("data", [])
        groups = sheet_def.get("groups", [])
        merge_rules = sheet_def.get("merge_rules", [])

        ws.title = title

        # Add group headers if defined
        if groups:
            current_row = self._add_group_headers(ws, groups, current_row)

        # Add column headers
        current_row = self._add_headers(ws, headers, current_row)

        # Add data rows
        if data:
            current_row = self._add_data_rows(ws, data, field_names, current_row, merge_rules)

        # Auto-size columns
        self._auto_size_columns(ws)

    def _add_group_headers(self, ws, groups, start_row):
        """
        Add grouped headers (multi-level headers).

        Args:
            ws: Worksheet object
            groups: List of group definitions
            start_row: Starting row number

        Returns:
            int: Next row number
        """
        col = 1
        for group in groups:
            title = group.get("title", "")
            span = group.get("span", 1)

            # Write group title
            cell = ws.cell(row=start_row, column=col, value=title)
            cell.font = Font(bold=True, size=DEFAULT_HEADER_FONT_SIZE)
            cell.fill = PatternFill(
                start_color=DEFAULT_HEADER_BG_COLOR, end_color=DEFAULT_HEADER_BG_COLOR, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal=DEFAULT_HEADER_ALIGNMENT, vertical="center")
            cell.border = self._get_border()

            # Merge cells if span > 1
            if span > 1:
                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=start_row,
                    end_column=col + span - 1,
                )

            col += span

        return start_row + 1

    def _add_headers(self, ws, headers, start_row):
        """
        Add column headers.

        Args:
            ws: Worksheet object
            headers: List of header strings
            start_row: Starting row number

        Returns:
            int: Next row number
        """
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=DEFAULT_HEADER_FONT_BOLD, size=DEFAULT_HEADER_FONT_SIZE)
            cell.fill = PatternFill(
                start_color=DEFAULT_HEADER_BG_COLOR, end_color=DEFAULT_HEADER_BG_COLOR, fill_type="solid"
            )
            cell.alignment = Alignment(horizontal=DEFAULT_HEADER_ALIGNMENT, vertical="center")
            cell.border = self._get_border()

        return start_row + 1

    def _add_data_rows(self, ws, data, field_names, start_row, merge_rules):
        """
        Add data rows to worksheet.

        Args:
            ws: Worksheet object
            data: List of data dictionaries
            field_names: List of field names corresponding to columns
            start_row: Starting row number
            merge_rules: List of field names to merge vertically

        Returns:
            int: Next row number
        """
        if not data:
            return start_row

        # Track merge ranges for each merge rule field
        merge_ranges = {field: [] for field in merge_rules}
        prev_values = dict.fromkeys(merge_rules)
        merge_start = dict.fromkeys(merge_rules, start_row)

        # Track rows processed in this batch for progress callback
        rows_before = self.total_rows_processed

        # Write data and track merge ranges
        self._write_data_rows(ws, data, field_names, start_row, merge_rules, merge_ranges, prev_values, merge_start)

        # Record final merge ranges
        self._finalize_merge_ranges(data, field_names, start_row, merge_rules, merge_ranges, merge_start)

        # Apply merges
        self._apply_cell_merges(ws, merge_ranges)

        # Call progress callback for any remaining rows not yet reported
        rows_processed_in_batch = self.total_rows_processed - rows_before
        remaining_rows = rows_processed_in_batch % self.chunk_size
        if self.progress_callback and remaining_rows > 0:
            self.progress_callback(remaining_rows)

        return start_row + len(data)

    def _write_data_rows(self, ws, data, field_names, start_row, merge_rules, merge_ranges, prev_values, merge_start):
        """Write data to cells and track values for merging."""
        for row_idx, row_data in enumerate(data):
            current_row = start_row + row_idx

            for col, field_name in enumerate(field_names, start=1):
                value = row_data.get(field_name, "")
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = self._get_border()

                # Track merge ranges for this field if needed
                if field_name in merge_rules:
                    self._track_merge_value(
                        field_name, value, current_row, col, merge_ranges, prev_values, merge_start
                    )

            # Update progress tracking
            self.total_rows_processed += 1
            if self.progress_callback and self.total_rows_processed % self.chunk_size == 0:
                self.progress_callback(self.chunk_size)

            # Artificial per-row delay for testing queue behavior.
            # Controlled by EXPORTER_ROW_DELAY_SECONDS env var (0 = disabled).
            row_delay = getattr(settings, "EXPORTER_ROW_DELAY_SECONDS", 0)
            if row_delay and row_delay > 0:
                logger.debug(f"Delaying {row_delay}s after writing row {self.total_rows_processed} (test mode)")
                time.sleep(row_delay)

    def _track_merge_value(self, field_name, current_value, current_row, col, merge_ranges, prev_values, merge_start):
        """Track value changes for cell merging."""
        # If value changed, record merge range for previous value
        if prev_values[field_name] is not None and current_value != prev_values[field_name]:
            if current_row - merge_start[field_name] > 1:
                merge_ranges[field_name].append((merge_start[field_name], current_row - 1, col))
            merge_start[field_name] = current_row

        prev_values[field_name] = current_value

    def _finalize_merge_ranges(self, data, field_names, start_row, merge_rules, merge_ranges, merge_start):
        """Record final merge ranges after all data is written."""
        final_row = start_row + len(data) - 1
        for field_name in merge_rules:
            col = field_names.index(field_name) + 1
            if final_row - merge_start[field_name] >= 1:
                merge_ranges[field_name].append((merge_start[field_name], final_row, col))

    def _apply_cell_merges(self, ws, merge_ranges):
        """Apply cell merges to worksheet."""
        for field_name, ranges in merge_ranges.items():
            for start_r, end_r, col in ranges:
                if end_r > start_r:
                    ws.merge_cells(start_row=start_r, start_column=col, end_row=end_r, end_column=col)
                    # Center align merged cells
                    ws.cell(row=start_r, column=col).alignment = Alignment(horizontal="center", vertical="center")

    def _get_border(self):
        """
        Get default border style.

        Returns:
            Border: Border style object
        """
        thin_border = Side(style="thin", color="000000")
        return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

    def _auto_size_columns(self, ws):
        """
        Auto-size columns based on content.

        Args:
            ws: Worksheet object
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception as e:
                    logger.debug(f"Skipping cell due to error: {e}")

            # Set column width (add padding)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _adjust_sheet_title(self, ws: Worksheet) -> None:
        """
        Adjust sheet title to fit content, especially for template that has logo image.

        NOTE: Use for export with template only.
        """
        default_title_col_letter = "B"  # NOTE: Hardcoded for now, since the template is BE self-prepared.
        default_title_row_inx = 2  # NOTE: Hardcoded for now, since the template is BE self-prepared.
        ws.column_dimensions[default_title_col_letter].width = 20
        ws.row_dimensions[default_title_row_inx].height = 25
