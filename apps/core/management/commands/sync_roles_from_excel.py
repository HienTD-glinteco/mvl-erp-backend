from __future__ import annotations

import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from apps.core.models import Permission, Role

MASTER_SHEET_NAME = "CORE_PERMISSION"
PERMISSION_SHEET_PREFIX = "S_"
ROLE_SHEET_PREFIX = "R-"
REQUIRED_PERMISSION_FIELDS = ("name", "description", "module", "submodule")
HEADER_ALIASES = {
    "code": "code",
    "ten": "name",
    "mo_ta": "description",
    "mota": "description",
    "module": "module",
    "submodule": "submodule",
    "nhom": "group",
}

SOURCE_CONTEXT_KEY = "__sources"


@dataclass
class RolePayload:
    """Container for role data parsed from Excel."""

    code: str
    name: str
    sheet_name: str
    permission_codes: List[str] = field(default_factory=list)


class WorkbookPermissionExtractor:
    """
    Parse Excel workbook into permission metadata and role definitions.
    run: python manage.py sync_roles_from_excel --file_path apps/core/fixtures/core_permission.xlsx  --verbose
    """

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.permission_catalog: Dict[str, Dict[str, Any]] = {}
        self.roles: List[RolePayload] = []

    def parse(self) -> Dict[str, Any]:
        workbook = self._load_workbook()
        try:
            self._ingest_master_sheet(workbook)
            for sheet in workbook.worksheets:
                title = (sheet.title or "").strip()
                if not title:
                    continue
                normalized_title = title.upper()
                if normalized_title == MASTER_SHEET_NAME:
                    continue
                if normalized_title.startswith(PERMISSION_SHEET_PREFIX):
                    self._ingest_permission_sheet(sheet)
                elif self._is_role_sheet(normalized_title):
                    self._ingest_role_sheet(sheet)
        finally:
            workbook.close()

        if not self.permission_catalog:
            raise CommandError("No permissions were found in the workbook.")
        if not self.roles:
            raise CommandError("No role sheets (prefixed with 'R-') were found in the workbook.")

        return {
            "permissions": self.permission_catalog,
            "roles": self.roles,
        }

    def _load_workbook(self):
        if not self.file_path.exists():
            raise CommandError(f"Workbook not found: {self.file_path}")
        try:
            return load_workbook(self.file_path, data_only=True)
        except InvalidFileException as exc:
            raise CommandError(f"Invalid Excel file: {self.file_path}") from exc
        except Exception as exc:  # pragma: no cover - best effort guard
            raise CommandError(f"Failed to load Excel file: {exc}") from exc

    def _ingest_master_sheet(self, workbook: Workbook) -> None:
        target_sheet: Optional[Worksheet] = None
        for sheet in workbook.worksheets:
            if (sheet.title or "").strip().upper() == MASTER_SHEET_NAME:
                target_sheet = sheet
                break
        if target_sheet is None:
            raise CommandError("Sheet 'core_permission' is required in the workbook.")
        self._ingest_permission_sheet(target_sheet)

    def _ingest_permission_sheet(self, sheet: Worksheet):
        for row, row_number in self._iter_data_rows(sheet):
            self._store_permission_metadata(row, sheet.title or "", row_number)

    def _ingest_role_sheet(self, sheet: Worksheet):
        role_code, role_name = self._derive_role_code_and_name(sheet.title)
        if not role_code or not role_name:
            raise CommandError(f"Role sheet '{sheet.title}' must follow format 'R-CODE-Name'.")

        codes: List[str] = []
        for row, row_number in self._iter_data_rows(sheet):
            code = self._store_permission_metadata(row, sheet.title or "", row_number)
            if code:
                codes.append(code)

        unique_codes = sorted(set(codes))
        if not unique_codes:
            raise CommandError(f"Role sheet '{sheet.title}' does not contain any permission codes.")

        self.roles.append(
            RolePayload(code=role_code, name=role_name, sheet_name=sheet.title, permission_codes=unique_codes)
        )

    def _iter_data_rows(self, sheet: Worksheet) -> Iterable[tuple[Dict[str, Any], int]]:
        tables = list(sheet.tables.values())
        if tables:
            for table in tables:
                yield from self._iter_table(sheet, table.ref)
            return

        max_row = sheet.max_row or 0
        max_col = sheet.max_column or 0
        if max_row < 2 or max_col < 1:
            return
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        yield from self._iter_table(sheet, ref)

    def _iter_table(self, sheet: Worksheet, cell_range: str) -> Iterable[tuple[Dict[str, Any], int]]:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        header: Optional[list[str | None]] = None
        for absolute_row_index, row in enumerate(
            sheet.iter_rows(
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
                values_only=True,
            ),
            start=min_row,
        ):
            if header is None:
                header = self._normalize_headers(row)
                continue
            row_data = self._build_row_dict(header, row)
            if row_data:
                yield row_data, absolute_row_index

    def _normalize_headers(self, header_row: Iterable[Any]) -> list[str | None]:
        normalized: list[str | None] = []
        for cell in header_row:
            normalized.append(self._canonical_header(cell))
        return normalized

    def _canonical_header(self, raw_value: Any) -> Optional[str]:
        if raw_value is None:
            return None
        normalized = self._strip_accents(str(raw_value))
        normalized = normalized.replace("-", "_")
        normalized = "_".join(part for part in normalized.split() if part)
        normalized = normalized.lower()
        return HEADER_ALIASES.get(normalized, normalized) or None

    def _build_row_dict(self, headers: list[str | None], row: Iterable[Any]) -> dict[str, None | Any]:
        data: dict[str, None | Any] = {}
        for header, value in zip(headers, row, strict=False):
            if not header:
                continue
            if isinstance(value, str):
                cleaned = value.strip()
            elif value is None:
                cleaned = ""
            else:
                cleaned = str(value).strip()
            data[header] = cleaned
        return data

    def _store_permission_metadata(self, row: Dict[str, Any], sheet_name: str, row_number: int) -> Optional[str]:
        raw_code = row.get("code")
        code = self._normalize_code(raw_code)
        if not code:
            return None

        default_entry: Dict[str, Any] = {SOURCE_CONTEXT_KEY: []}
        default_entry.update(dict.fromkeys(REQUIRED_PERMISSION_FIELDS, ""))
        entry = self.permission_catalog.setdefault(code, default_entry)
        source_entries: List[Dict[str, Any]] = entry.setdefault(SOURCE_CONTEXT_KEY, [])
        source_entries.append({"sheet": sheet_name, "row": row_number})
        for _field in REQUIRED_PERMISSION_FIELDS:
            if entry.get(_field):
                continue
            candidate = row.get(_field)
            if candidate:
                entry[_field] = candidate
        return code

    def _normalize_code(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        code = str(value).strip()
        return code or None

    def _strip_accents(self, value: str) -> str:
        normalized = unicodedata.normalize("NFD", value)
        return "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")

    def _derive_role_code_and_name(self, sheet_title: str) -> tuple[str, str]:
        """
        Derive role code and name from sheet title.
        Format: R-CODE-Name -> code=CODE, name=Name
        """
        trimmed = (sheet_title or "").strip()
        if not trimmed.upper().startswith(ROLE_SHEET_PREFIX):
            return "", ""

        # Remove the R- prefix
        remainder = trimmed[len(ROLE_SHEET_PREFIX) :]

        # Split by the first dash to get code and name
        parts = remainder.split("-", 1)
        if len(parts) < 2:
            return "", ""

        role_code = parts[0].strip()
        role_name = parts[1].strip()

        return role_code, role_name

    def _is_role_sheet(self, normalized_title: str) -> bool:
        return normalized_title.startswith(ROLE_SHEET_PREFIX)


class Command(BaseCommand):
    help = "Sync permissions and roles defined in an Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file_path",
            type=str,
            help="Path to the Excel workbook (e.g. core_permission.xlsx).",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simulate the import without persisting changes.",
        )
        parser.add_argument(
            "--verbose",
            action="store_true",
            help="Enable verbose output.",
        )

    def handle(self, *args, **options):
        if options.get("file_path"):
            file_path = Path(options["file_path"]).expanduser()
        else:
            file_path = Path(__file__).resolve().parents[2] / "fixtures" / "core_permission.xlsx"
        self.verbose = options.get("verbose", False)
        self.dry_run = options.get("dry_run", False)

        extractor = WorkbookPermissionExtractor(file_path)
        workbook_data = extractor.parse()

        excel_permission_codes = set(workbook_data["permissions"].keys())
        self._print_permissions_not_in_excel(excel_permission_codes)

        with transaction.atomic():
            permission_stats, permission_objects = self._sync_permissions(workbook_data["permissions"])
            role_stats = self._sync_roles(workbook_data["roles"], permission_objects)
            if self.dry_run:
                transaction.set_rollback(True)

        self._print_summary(permission_stats, role_stats)

    def _sync_permissions(self, permission_catalog: Dict[str, Dict[str, Any]]):
        stats = {"processed": 0, "mismatched": 0}
        permission_objects: Dict[str, Permission] = {}
        mismatch_entries: List[Dict[str, str]] = []

        for code, metadata in permission_catalog.items():
            stats["processed"] += 1
            source_context: Optional[List[Dict[str, Any]]] = metadata.get("__sources")
            location = ""
            try:
                permission = Permission.objects.get(code=code)
                if source_context:
                    origin = source_context[0]
                    sheet_label = (origin.get("sheet") or "").strip() or "Unknown sheet"
                    row_label = origin.get("row")
                    location = (
                        f" (sheet '{sheet_label}', row {row_label})" if row_label else f" (sheet '{sheet_label}')"
                    )
            except Permission.DoesNotExist:
                if source_context:
                    origin = source_context[0]
                    sheet_label = (origin.get("sheet") or "").strip() or "Unknown sheet"
                    row_label = origin.get("row")
                    location = (
                        f" (sheet '{sheet_label}', row {row_label})" if row_label else f" (sheet '{sheet_label}')"
                    )
                raise CommandError(f"Permission '{code}' does not exist in the database{location}.")

            permission_objects[code] = permission
            mismatches = {}
            for _field in ("name", "description"):
                workbook_value = str(metadata.get(_field) or "").strip()
                db_value = getattr(permission, _field, "")
                if workbook_value and workbook_value != db_value:
                    mismatches[_field] = db_value
            if mismatches:
                stats["mismatched"] += 1
                mismatch_entries.append(
                    {
                        "code": code,
                        "name": mismatches.get("name", permission.name),
                        "description": mismatches.get("description", permission.description),
                        "location": location,
                    }
                )

        if mismatch_entries:
            self._print_mismatch_table(mismatch_entries)

        return stats, permission_objects

    def _sync_roles(self, roles: List[RolePayload], permission_objects: Dict[str, Permission]):
        stats = {"processed": 0, "created": 0, "updated": 0, "permissions_updated": 0}

        for role_payload in roles:
            stats["processed"] += 1

            # Try to get existing role by code first
            try:
                role = Role.objects.get(code=role_payload.code)
                created = False
            except Role.DoesNotExist:
                # Check if a role with this name already exists (name is unique)
                existing_by_name = Role.objects.filter(name=role_payload.name).first()
                if existing_by_name:
                    raise CommandError(
                        f"Cannot create role '{role_payload.code}' with name '{role_payload.name}': "
                        f"a role with this name already exists (code: '{existing_by_name.code}')."
                    )
                role = Role.objects.create(
                    code=role_payload.code,
                    name=role_payload.name,
                    description="",
                    is_system_role=True,
                )
                created = True

            if created:
                stats["created"] += 1
                self._log(
                    "debug",
                    f"Created role {role_payload.code} - {role_payload.name} (sheet: {role_payload.sheet_name})",
                )
            else:
                updates: List[str] = []
                if role.name != role_payload.name:
                    # Check if new name conflicts with another role
                    existing_by_name = (
                        Role.objects.filter(name=role_payload.name).exclude(code=role_payload.code).first()
                    )
                    if existing_by_name:
                        raise CommandError(
                            f"Cannot update role '{role_payload.code}' name to '{role_payload.name}': "
                            f"a role with this name already exists (code: '{existing_by_name.code}')."
                        )
                    role.name = role_payload.name
                    updates.append("name")
                if role.is_system_role is not True:
                    role.is_system_role = True
                    updates.append("is_system_role")
                if updates:
                    role.save(update_fields=updates)
                    stats["updated"] += 1

            target_permissions = [
                permission_objects[code] for code in role_payload.permission_codes if code in permission_objects
            ]
            previous_count = role.permissions.count()
            role.permissions.set(target_permissions)
            if previous_count != len(target_permissions):
                stats["permissions_updated"] += 1

        return stats

    def _log(self, level: str, message: str):
        if level == "debug" and not getattr(self, "verbose", False):
            return
        if level == "warning":
            self.stdout.write(self.style.WARNING(f"⚠ {message}"))
        elif level == "error":
            self.stdout.write(self.style.ERROR(f"✗ {message}"))
        else:
            self.stdout.write(f"  → {message}")

    def _print_mismatch_table(self, mismatch_entries: List[Dict[str, str]]):
        """Print permission metadata mismatches in table format (DB values)."""
        self.stdout.write("")
        self.stdout.write(self.style.WARNING(f"Permission metadata mismatches ({len(mismatch_entries)} items):"))
        self.stdout.write("Permissions with code in Excel but name/description differ from DB values:")
        self.stdout.write("")
        self.stdout.write(f"{'code':<40} | {'name':<40} | {'description'} | {'location':<40}")
        self.stdout.write("-" * 120)
        for entry in mismatch_entries:
            code = entry["code"]
            name = entry["name"] if entry["name"] else ""
            desc = entry["description"] if entry["description"] else ""
            location = entry["location"] if entry.get("location") else ""
            self.stdout.write(f"{code:<40} | {name:<40} | {desc} | {location:<40}")
        self.stdout.write("")

    def _print_permissions_not_in_excel(self, excel_permission_codes: set):
        """Print permissions in DB but not in Excel in table format for copy/paste."""
        db_codes = set(Permission.objects.values_list("code", flat=True))
        missing_from_excel = db_codes - excel_permission_codes
        if not missing_from_excel:
            return

        missing_permissions = Permission.objects.filter(code__in=missing_from_excel).order_by(
            "module", "submodule", "code"
        )

        self.stdout.write("")
        self.stdout.write(self.style.WARNING(f"Missing permissions ({len(missing_from_excel)} items):"))
        self.stdout.write("Permissions in DB but code NOT declared in core_permission.xlsx:")
        self.stdout.write("")
        self.stdout.write(f"{'code':<40} | {'name':<30} | {'description':<30} | {'module':<15} | {'submodule'}")
        self.stdout.write("-" * 140)
        for perm in missing_permissions:
            code = perm.code
            name = perm.name or ""
            desc = perm.description or ""
            module = perm.module or ""
            submodule = perm.submodule or ""
            self.stdout.write(f"{code:<40} | {name:<30} | {desc:<30} | {module:<15} | {submodule}")
        self.stdout.write("")

    def _print_summary(self, permission_stats: Dict[str, int], role_stats: Dict[str, int]):
        self.stdout.write("")
        self.stdout.write(self.style.SUCCESS("=" * 60))
        self.stdout.write(self.style.SUCCESS("EXCEL ROLE SYNC SUMMARY"))
        self.stdout.write(self.style.SUCCESS("=" * 60))

        self.stdout.write(f"Permissions checked: {permission_stats['processed']}")
        self.stdout.write(f"  Name/description mismatches: {permission_stats['mismatched']}")

        self.stdout.write("")
        self.stdout.write(f"Roles processed: {role_stats['processed']}")
        self.stdout.write(f"  Created: {role_stats['created']}")
        self.stdout.write(f"  Updated metadata: {role_stats['updated']}")
        self.stdout.write(f"  Permissions refreshed: {role_stats['permissions_updated']}")

        self.stdout.write(self.style.SUCCESS("=" * 60))
