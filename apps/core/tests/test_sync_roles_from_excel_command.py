from pathlib import Path

import pytest
from django.core.management import CommandError, call_command
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from apps.core.models import Permission, Role


def _add_table(sheet, name: str, ref: str):
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)


def _build_sample_workbook(file_path: Path):
    workbook = Workbook()
    core_sheet = workbook.active
    core_sheet.title = "core_permission"
    core_sheet.append(["code", "Tên", "Mô tả", "module", "submodule"])
    core_sheet.append(["permission.create", "Create Permission", "Create Permission", "Core", "Permission"])
    core_sheet.append(["permission.list", "List Permission", "List Permission", "Core", "Permission"])
    _add_table(core_sheet, "CorePermissions", f"A1:E{core_sheet.max_row}")

    set_sheet = workbook.create_sheet("S_EMPLOYEE")
    set_sheet.append(["code", "Tên", "Mô tả", "module", "submodule", "Nhóm"])
    set_sheet.append(
        ["permission.create", "Create Permission", "Create Permission", "Core", "Permission", "S_EMPLOYEE"]
    )
    set_sheet.append(["permission.list", "List Permission", "List Permission", "Core", "Permission", "S_EMPLOYEE"])
    _add_table(set_sheet, "S_EMPLOYEE_TABLE", f"A1:F{set_sheet.max_row}")

    role_sheet = workbook.create_sheet("R-Employee")
    role_sheet.append(["code", "Tên", "Mô tả", "module", "submodule", "Nhóm"])
    role_sheet.append(
        ["permission.create", "Create Permission", "Create Permission", "Core", "Permission", "S_EMPLOYEE"]
    )
    role_sheet.append(["permission.list", "List Permission", "List Permission", "Core", "Permission", "S_EMPLOYEE"])

    workbook.save(file_path)
    workbook.close()


def _create_permission(code: str, name: str, description: str):
    Permission.objects.create(
        code=code,
        name=name,
        description=description,
        module="Core",
        submodule="Permission",
    )


@pytest.mark.django_db
def test_sync_roles_from_excel_attaches_permissions_and_roles(tmp_path):
    file_path = tmp_path / "permissions.xlsx"
    _build_sample_workbook(file_path)

    _create_permission("permission.create", "Create Permission", "Create Permission")
    _create_permission("permission.list", "List Permission", "List Permission")

    call_command("sync_roles_from_excel", "--file_path", str(file_path))

    assert Permission.objects.count() == 2
    assert Role.objects.filter(name="Employee").exists()

    role = Role.objects.get(name="Employee")
    assert role.is_system_role is True
    assert set(role.permissions.values_list("code", flat=True)) == {"permission.create", "permission.list"}


@pytest.mark.django_db
def test_sync_roles_from_excel_dry_run(tmp_path):
    file_path = tmp_path / "permissions_dry_run.xlsx"
    _build_sample_workbook(file_path)

    _create_permission("permission.create", "Create Permission", "Create Permission")
    _create_permission("permission.list", "List Permission", "List Permission")

    call_command("sync_roles_from_excel", "--file_path", str(file_path), "--dry-run")

    assert Permission.objects.count() == 2
    assert Role.objects.count() == 0


@pytest.mark.django_db
def test_sync_roles_from_excel_fails_for_unknown_permission(tmp_path):
    file_path = tmp_path / "permissions_unknown.xlsx"
    _build_sample_workbook(file_path)

    with pytest.raises(CommandError) as excinfo:
        call_command("sync_roles_from_excel", "--file_path", str(file_path))

    message = str(excinfo.value)
    assert "does not exist in the database" in message
    assert "sheet 'core_permission'" in message
    assert "row 2" in message


@pytest.mark.django_db
def test_sync_roles_from_excel_warns_on_metadata_mismatch(tmp_path, capsys):
    file_path = tmp_path / "permissions_mismatch.xlsx"
    _build_sample_workbook(file_path)

    _create_permission("permission.create", "Different Name", "Different Description")
    _create_permission("permission.list", "List Permission", "List Permission")

    call_command("sync_roles_from_excel", "--file_path", str(file_path))

    captured = capsys.readouterr()
    assert "metadata mismatch" in captured.out
