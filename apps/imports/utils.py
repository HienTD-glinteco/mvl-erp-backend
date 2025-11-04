"""Utility functions for import operations."""

import csv
import io
import logging
import os
import tempfile
from typing import Iterator, Optional, TextIO

import openpyxl
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from apps.files.models import FileModel

logger = logging.getLogger(__name__)


class StreamingReader:
    """Base class for streaming file readers."""

    def __init__(self, file_path: str):
        """
        Initialize reader.

        Args:
            file_path: S3 path or local file path
        """
        self.file_path = file_path
        self.file_stream = None

    def __enter__(self):
        """Open the file stream."""
        self.file_stream = default_storage.open(self.file_path, "rb")
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Close the file stream."""
        if self.file_stream:
            self.file_stream.close()

    def read_rows(self, skip_rows: int = 1) -> Iterator[list]:
        """
        Read rows from file.

        Args:
            skip_rows: Number of header rows to skip

        Yields:
            list: Row data as list
        """
        raise NotImplementedError


class CSVStreamingReader(StreamingReader):
    """Streaming reader for CSV files."""

    def read_rows(self, skip_rows: int = 1) -> Iterator[list]:
        """
        Read rows from CSV file.

        Args:
            skip_rows: Number of header rows to skip

        Yields:
            list: Row data as list
        """
        # Wrap binary stream in TextIOWrapper for csv.reader
        if self.file_stream is None:
            raise ValueError("file_stream is not initialized")

        text_stream = io.TextIOWrapper(self.file_stream, encoding="utf-8-sig")
        reader = csv.reader(text_stream)

        # Skip header rows
        for _ in range(skip_rows):
            try:
                next(reader)
            except StopIteration:
                return

        # Yield data rows
        for row in reader:
            yield row


class XLSXStreamingReader(StreamingReader):
    """Streaming reader for XLSX files."""

    def read_rows(self, skip_rows: int = 1) -> Iterator[list]:
        """
        Read rows from XLSX file.

        Args:
            skip_rows: Number of header rows to skip

        Yields:
            list: Row data as list
        """
        # Load workbook in read-only mode for streaming
        workbook = openpyxl.load_workbook(self.file_stream, read_only=True, data_only=True)
        sheet = workbook.active

        # Iterate rows
        row_index = 0
        for row in sheet.iter_rows(values_only=True):
            row_index += 1

            # Skip header rows
            if row_index <= skip_rows:
                continue

            # Convert row tuple to list and yield
            yield list(row)

        workbook.close()


class StreamingWriter:
    """Base class for streaming result file writers."""

    def __init__(self, temp_dir: Optional[str] = None):
        """
        Initialize writer.

        Args:
            temp_dir: Temporary directory for file (None = system temp)
        """
        self.temp_dir = temp_dir or tempfile.gettempdir()
        self.temp_file: Optional[TextIO] = None
        self.file_path: Optional[str] = None

    def __enter__(self):
        """Create temporary file."""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Clean up temporary file."""
        self.close()

    def write_header(self, headers: list) -> None:
        """
        Write header row.

        Args:
            headers: List of header values
        """
        raise NotImplementedError

    def write_row(self, row: list) -> None:
        """
        Write data row.

        Args:
            row: List of cell values
        """
        raise NotImplementedError

    def close(self) -> None:
        """Close the file."""
        raise NotImplementedError

    def get_file_path(self) -> str:
        """
        Get the temporary file path.

        Returns:
            str: Path to temporary file
        """
        if self.file_path is None:
            raise ValueError("file_path must be set before calling get_file_path()")
        return self.file_path


class CSVStreamingWriter(StreamingWriter):
    """Streaming writer for CSV files."""

    def __init__(self, filename: str, temp_dir: Optional[str] = None):
        """
        Initialize CSV writer.

        Args:
            filename: Name of the file
            temp_dir: Temporary directory
        """
        super().__init__(temp_dir)
        self.file_path = os.path.join(self.temp_dir, filename)
        self.temp_file = open(self.file_path, "w", newline="", encoding="utf-8")
        self.writer = csv.writer(self.temp_file)

    def write_header(self, headers: list) -> None:
        """Write CSV header."""
        self.writer.writerow(headers)

    def write_row(self, row: list) -> None:
        """Write CSV row."""
        self.writer.writerow(row)

    def close(self) -> None:
        """Close CSV file."""
        if self.temp_file:
            self.temp_file.close()


class XLSXStreamingWriter(StreamingWriter):
    """Streaming writer for XLSX files."""

    def __init__(self, filename: str, temp_dir: Optional[str] = None):
        """
        Initialize XLSX writer.

        Args:
            filename: Name of the file
            temp_dir: Temporary directory
        """
        super().__init__(temp_dir)
        self.file_path = os.path.join(self.temp_dir, filename)
        self.workbook = openpyxl.Workbook(write_only=True)
        self.sheet = self.workbook.create_sheet()

    def write_header(self, headers: list) -> None:
        """Write XLSX header."""
        self.sheet.append(headers)

    def write_row(self, row: list) -> None:
        """Write XLSX row."""
        self.sheet.append(row)

    def close(self) -> None:
        """Close XLSX file."""
        if self.workbook:
            self.workbook.save(self.file_path)
            self.workbook.close()


def get_streaming_reader(file_path: str, file_extension: str) -> StreamingReader:
    """
    Get appropriate streaming reader based on file extension.

    Args:
        file_path: S3 path or local file path
        file_extension: File extension (.csv or .xlsx)

    Returns:
        StreamingReader: Appropriate reader instance
    """
    ext = file_extension.lower()
    if ext in [".csv", ".txt"]:
        return CSVStreamingReader(file_path)
    elif ext in [".xlsx", ".xls"]:
        return XLSXStreamingReader(file_path)
    else:
        # Default to CSV for unknown types
        return CSVStreamingReader(file_path)


def get_streaming_writer(filename: str, output_format: str = "csv", temp_dir: Optional[str] = None) -> StreamingWriter:
    """
    Get appropriate streaming writer based on output format.

    Args:
        filename: Base filename without extension
        output_format: Output format ('csv' or 'xlsx')
        temp_dir: Temporary directory

    Returns:
        StreamingWriter: Appropriate writer instance
    """
    if output_format.lower() == "xlsx":
        return XLSXStreamingWriter(f"{filename}.xlsx", temp_dir)
    else:
        return CSVStreamingWriter(f"{filename}.csv", temp_dir)


def upload_result_file(
    local_file_path: str,
    s3_prefix: str,
    original_filename: str,
    purpose: str,
    uploaded_by=None,
) -> FileModel:
    """
    Upload result file to S3 and create FileModel record.

    Args:
        local_file_path: Local path to the file
        s3_prefix: S3 prefix for the file
        original_filename: Original filename
        purpose: File purpose
        uploaded_by: User who uploaded the file

    Returns:
        FileModel: Created FileModel instance
    """
    # Generate S3 path
    filename = os.path.basename(local_file_path)
    s3_path = f"{s3_prefix}{filename}"

    # Read file content
    with open(local_file_path, "rb") as f:
        file_content = f.read()

    # Save to S3
    saved_path = default_storage.save(s3_path, ContentFile(file_content))

    # Get file size
    file_size = os.path.getsize(local_file_path)

    # Create FileModel record
    file_model = FileModel.objects.create(
        purpose=purpose,
        file_name=original_filename,
        file_path=saved_path,
        size=file_size,
        is_confirmed=True,
        uploaded_by=uploaded_by,
    )

    return file_model


def count_total_rows(file_path: str, file_extension: str, skip_rows: int = 1) -> int:
    """
    Count total rows in a file.

    Args:
        file_path: S3 path or local file path
        file_extension: File extension
        skip_rows: Number of header rows to skip

    Returns:
        int: Total number of data rows
    """
    count = 0
    reader = get_streaming_reader(file_path, file_extension)
    with reader:
        for _ in reader.read_rows(skip_rows=skip_rows):
            count += 1
    return count


def read_headers(file_path: str, file_extension: str, header_row: int = 0) -> list:
    """
    Read header row from a file.

    Args:
        file_path: S3 path or local file path
        file_extension: File extension
        header_row: 0-based index of header row (default: 0 for first row)

    Returns:
        list: List of header values
    """
    headers = []

    with default_storage.open(file_path, "rb") as file_stream:
        ext = file_extension.lower()

        if ext in [".csv", ".txt"]:
            # CSV file
            text_stream = io.TextIOWrapper(file_stream, encoding="utf-8-sig")
            reader = csv.reader(text_stream)

            # Skip to header row
            for i, row in enumerate(reader):
                if i == header_row:
                    headers = row
                    break

        elif ext in [".xlsx", ".xls"]:
            # XLSX file
            workbook = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)
            sheet = workbook.active

            # Get header row (1-based in openpyxl)
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == header_row:
                    headers = [cell if cell is not None else "" for cell in row]
                    break

            workbook.close()

    return headers
