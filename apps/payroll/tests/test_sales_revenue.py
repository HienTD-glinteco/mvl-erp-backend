import json
from datetime import date
from io import BytesIO

import openpyxl
import pytest
from django.contrib.auth import get_user_model
from rest_framework import status

from apps.hrm.models import Employee
from apps.payroll.models import SalesRevenue

User = get_user_model()


@pytest.fixture
def sales_employee(employee):
    """Use the existing employee fixture as sales employee."""
    return employee


@pytest.fixture
def inactive_employee(branch, block, department, position):
    """Create an inactive employee for testing."""
    from datetime import date

    from apps.payroll.tests.conftest import random_code

    code = random_code()
    return Employee.objects.create(
        username=f"inactive{code}",
        email=f"inactive{code}@example.com",
        phone=f"09{random_code(length=8)}",
        citizen_id=f"{random_code(length=12)}",
        status="inactive",
        branch=branch,
        block=block,
        department=department,
        position=position,
        start_date=date.today(),
    )


@pytest.fixture
def sales_revenue(sales_employee, user):
    """Create a sales revenue record for testing."""
    return SalesRevenue.objects.create(
        employee=sales_employee,
        revenue=150000000,
        transaction_count=12,
        month=date(2025, 11, 1),
        created_by=user,
    )


@pytest.mark.django_db
class TestSalesRevenueModel:
    """Test SalesRevenue model."""

    def test_create_sales_revenue(self, sales_employee, user):
        """Test creating a sales revenue record."""
        # Arrange
        revenue = 150000000
        transaction_count = 12
        month = date(2025, 11, 1)

        # Act
        sales_revenue = SalesRevenue.objects.create(
            employee=sales_employee,
            revenue=revenue,
            transaction_count=transaction_count,
            month=month,
            created_by=user,
        )

        # Assert
        assert sales_revenue.id is not None
        assert sales_revenue.code.startswith("SR-202511-")
        assert sales_revenue.employee == sales_employee
        assert sales_revenue.revenue == revenue
        assert sales_revenue.transaction_count == transaction_count
        assert sales_revenue.month == month
        assert sales_revenue.status == SalesRevenue.SalesRevenueStatus.NOT_CALCULATED
        assert sales_revenue.created_by == user

    def test_auto_generate_code(self, sales_employee, user):
        """Test auto-generation of code."""
        # Arrange & Act
        sr1 = SalesRevenue.objects.create(
            employee=sales_employee,
            revenue=100000000,
            transaction_count=10,
            month=date(2025, 11, 1),
            created_by=user,
        )
        sr2 = SalesRevenue.objects.create(
            employee=sales_employee,
            revenue=200000000,
            transaction_count=20,
            month=date(2025, 12, 1),
            created_by=user,
        )

        # Assert
        assert sr1.code == "SR-202511-0001"
        assert sr2.code == "SR-202512-0001"

    def test_month_normalization(self, sales_employee, user):
        """Test month is normalized to first day of month."""
        # Arrange & Act
        sales_revenue = SalesRevenue.objects.create(
            employee=sales_employee,
            revenue=100000000,
            transaction_count=10,
            month=date(2025, 11, 15),
            created_by=user,
        )

        # Assert
        assert sales_revenue.month == date(2025, 11, 1)

    def test_unique_employee_month(self, sales_employee, user):
        """Test unique constraint on employee and month."""
        # Arrange
        SalesRevenue.objects.create(
            employee=sales_employee,
            revenue=100000000,
            transaction_count=10,
            month=date(2025, 11, 1),
            created_by=user,
        )

        # Act & Assert
        with pytest.raises(Exception):
            SalesRevenue.objects.create(
                employee=sales_employee,
                revenue=200000000,
                transaction_count=20,
                month=date(2025, 11, 1),
                created_by=user,
            )

    def test_reset_status_to_not_calculated(self, sales_revenue):
        """Test resetting status to NOT_CALCULATED."""
        # Arrange
        sales_revenue.status = SalesRevenue.SalesRevenueStatus.CALCULATED
        sales_revenue.save(update_fields=["status"])

        # Act
        sales_revenue.reset_status_to_not_calculated()

        # Assert
        sales_revenue.refresh_from_db()
        assert sales_revenue.status == SalesRevenue.SalesRevenueStatus.NOT_CALCULATED

    def test_colored_status(self, sales_revenue):
        """Test colored status property."""
        # Act
        colored = sales_revenue.colored_status

        # Assert
        assert colored is not None
        assert "value" in colored
        assert "variant" in colored

    def test_str_representation(self, sales_revenue):
        """Test string representation."""
        # Act
        result = str(sales_revenue)

        # Assert
        assert sales_revenue.code in result
        # Employee username is used in string representation
        assert result is not None


@pytest.mark.django_db
class TestSalesRevenueSerializer:
    """Test SalesRevenueSerializer."""

    def test_serialize_sales_revenue(self, sales_revenue):
        """Test serializing a sales revenue record."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        # Act
        serializer = SalesRevenueSerializer(sales_revenue)

        # Assert
        assert serializer.data["id"] == sales_revenue.id
        assert serializer.data["code"] == sales_revenue.code
        assert serializer.data["revenue"] == sales_revenue.revenue
        assert serializer.data["transaction_count"] == sales_revenue.transaction_count
        assert serializer.data["month"] == "11/2025"
        assert serializer.data["status"] == SalesRevenue.SalesRevenueStatus.NOT_CALCULATED
        assert "employee" in serializer.data
        assert serializer.data["employee"]["id"] == sales_revenue.employee.id

    def test_deserialize_sales_revenue(self, sales_employee):
        """Test deserializing sales revenue data."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        data = {
            "employee_id": sales_employee.id,
            "revenue": 180000000,
            "transaction_count": 15,
            "month": "11/2025",
        }

        # Act
        serializer = SalesRevenueSerializer(data=data)

        # Assert
        assert serializer.is_valid(), serializer.errors
        assert serializer.validated_data["employee"] == sales_employee
        assert serializer.validated_data["revenue"] == 180000000
        assert serializer.validated_data["transaction_count"] == 15
        assert serializer.validated_data["month"] == date(2025, 11, 1)

    def test_validate_employee_must_be_active(self, inactive_employee):
        """Test validation fails for inactive employee."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        data = {
            "employee_id": inactive_employee.id,
            "revenue": 100000000,
            "transaction_count": 10,
            "month": "11/2025",
        }

        # Act
        serializer = SalesRevenueSerializer(data=data)

        # Assert
        assert not serializer.is_valid()
        assert "employee_id" in serializer.errors

    def test_validate_revenue_required(self, sales_employee):
        """Test revenue is required."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        data = {
            "employee_id": sales_employee.id,
            "transaction_count": 10,
            "month": "11/2025",
        }

        # Act
        serializer = SalesRevenueSerializer(data=data)

        # Assert
        assert not serializer.is_valid()
        assert "revenue" in serializer.errors

    def test_validate_revenue_non_negative(self, sales_employee):
        """Test revenue must be non-negative."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        data = {
            "employee_id": sales_employee.id,
            "revenue": -100,
            "transaction_count": 10,
            "month": "11/2025",
        }

        # Act
        serializer = SalesRevenueSerializer(data=data)

        # Assert
        assert not serializer.is_valid()
        assert "revenue" in serializer.errors

    def test_validate_transaction_count_non_negative(self, sales_employee):
        """Test transaction count must be non-negative."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        data = {
            "employee_id": sales_employee.id,
            "revenue": 100000000,
            "transaction_count": -5,
            "month": "11/2025",
        }

        # Act
        serializer = SalesRevenueSerializer(data=data)

        # Assert
        assert not serializer.is_valid()
        assert "transaction_count" in serializer.errors

    def test_validate_month_format(self, sales_employee):
        """Test month format validation."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        data = {
            "employee_id": sales_employee.id,
            "revenue": 100000000,
            "transaction_count": 10,
            "month": "2025-11",
        }

        # Act
        serializer = SalesRevenueSerializer(data=data)

        # Assert
        assert not serializer.is_valid()
        assert "month" in serializer.errors

    def test_validate_duplicate_employee_month(self, sales_revenue):
        """Test duplicate validation."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        data = {
            "employee_id": sales_revenue.employee.id,
            "revenue": 200000000,
            "transaction_count": 20,
            "month": "11/2025",
        }

        # Act
        serializer = SalesRevenueSerializer(data=data)

        # Assert
        assert not serializer.is_valid()
        assert "non_field_errors" in serializer.errors

    def test_update_resets_status(self, sales_revenue):
        """Test update resets status to NOT_CALCULATED."""
        # Arrange
        from apps.payroll.api.serializers import SalesRevenueSerializer

        sales_revenue.status = SalesRevenue.SalesRevenueStatus.CALCULATED
        sales_revenue.save(update_fields=["status"])

        data = {
            "employee_id": sales_revenue.employee.id,
            "revenue": 200000000,
            "transaction_count": 20,
            "month": "11/2025",
        }

        # Act
        serializer = SalesRevenueSerializer(sales_revenue, data=data)
        assert serializer.is_valid()
        updated = serializer.save()

        # Assert
        assert updated.status == SalesRevenue.SalesRevenueStatus.NOT_CALCULATED


@pytest.mark.django_db
class TestSalesRevenueViewSet:
    """Test SalesRevenueViewSet."""

    def test_list_sales_revenues(self, api_client, superuser, sales_revenue):
        """Test listing sales revenues."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        # Act
        response = api_client.get("/api/payroll/sales-revenues/")

        # Assert
        assert response.status_code == status.HTTP_200_OK
        data = json.loads(response.content)
        assert data["success"] is True
        assert data["data"]["count"] == 1
        assert len(data["data"]["results"]) == 1
        assert data["data"]["results"][0]["code"] == sales_revenue.code

    def test_retrieve_sales_revenue(self, api_client, superuser, sales_revenue):
        """Test retrieving a sales revenue."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        # Act
        response = api_client.get(f"/api/payroll/sales-revenues/{sales_revenue.id}/")

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["code"] == sales_revenue.code
        assert response.data["revenue"] == sales_revenue.revenue

    def test_create_sales_revenue(self, api_client, superuser, sales_employee):
        """Test creating a sales revenue."""
        # Arrange
        api_client.force_authenticate(user=superuser)
        data = {
            "employee_id": sales_employee.id,
            "revenue": 180000000,
            "transaction_count": 15,
            "month": "12/2025",
        }

        # Act
        response = api_client.post("/api/payroll/sales-revenues/", data)

        # Assert
        assert response.status_code == status.HTTP_201_CREATED
        assert response.data["code"].startswith("SR-202512-")
        assert response.data["revenue"] == 180000000
        assert response.data["status"] == SalesRevenue.SalesRevenueStatus.NOT_CALCULATED
        assert SalesRevenue.objects.filter(employee=sales_employee, month=date(2025, 12, 1)).exists()

    def test_update_sales_revenue(self, api_client, superuser, sales_revenue):
        """Test updating a sales revenue."""
        # Arrange
        api_client.force_authenticate(user=superuser)
        data = {
            "employee_id": sales_revenue.employee.id,
            "revenue": 200000000,
            "transaction_count": 20,
            "month": "11/2025",
        }

        # Act
        response = api_client.put(f"/api/payroll/sales-revenues/{sales_revenue.id}/", data)

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["revenue"] == 200000000
        assert response.data["transaction_count"] == 20
        assert response.data["status"] == SalesRevenue.SalesRevenueStatus.NOT_CALCULATED

    def test_partial_update_sales_revenue(self, api_client, superuser, sales_revenue):
        """Test partially updating a sales revenue."""
        # Arrange
        api_client.force_authenticate(user=superuser)
        data = {"revenue": 175000000}

        # Act
        response = api_client.patch(f"/api/payroll/sales-revenues/{sales_revenue.id}/", data)

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["revenue"] == 175000000
        assert response.data["transaction_count"] == sales_revenue.transaction_count

    def test_delete_sales_revenue(self, api_client, superuser, sales_revenue):
        """Test deleting a sales revenue."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        # Act
        response = api_client.delete(f"/api/payroll/sales-revenues/{sales_revenue.id}/")

        # Assert
        assert response.status_code == status.HTTP_204_NO_CONTENT
        assert not SalesRevenue.objects.filter(id=sales_revenue.id).exists()

    def test_filter_by_month(self, api_client, superuser, sales_employee, user):
        """Test filtering by month."""
        # Arrange
        api_client.force_authenticate(user=superuser)
        SalesRevenue.objects.create(
            employee=sales_employee,
            revenue=100000000,
            transaction_count=10,
            month=date(2025, 11, 1),
            created_by=user,
        )
        SalesRevenue.objects.create(
            employee=sales_employee,
            revenue=200000000,
            transaction_count=20,
            month=date(2025, 12, 1),
            created_by=user,
        )

        # Act
        response = api_client.get("/api/payroll/sales-revenues/?month=11/2025")

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] == 1
        assert response.data["results"][0]["month"] == "11/2025"

    def test_filter_by_status(self, api_client, superuser, sales_revenue):
        """Test filtering by status."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        # Act
        response = api_client.get(
            f"/api/payroll/sales-revenues/?status={SalesRevenue.SalesRevenueStatus.NOT_CALCULATED}"
        )

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] >= 1

    def test_search_by_employee_code(self, api_client, superuser, sales_revenue):
        """Test searching by employee code."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        # Act
        response = api_client.get(f"/api/payroll/sales-revenues/?search={sales_revenue.employee.code}")

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] >= 1

    def test_search_by_employee_name(self, api_client, superuser, sales_revenue):
        """Test searching by employee name."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        # Act
        response = api_client.get(f"/api/payroll/sales-revenues/?search={sales_revenue.employee.username}")

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["count"] >= 1


@pytest.mark.django_db
class TestSalesRevenueUpload:
    """Test sales revenue upload functionality."""

    def test_upload_sales_revenue_success(self, api_client, superuser, sales_employee):
        """Test successful upload of sales revenue data."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Mã nhân viên", "Họ tên", "Doanh Số", "Số giao dịch", "Thời gian"])
        ws.append(["Ký tự (20)", "Ký tự (100)", "Số (20)", "Số (10)", "MM/YYYY"])
        ws.append([sales_employee.code, sales_employee.username, 150000000, 12, "11/2025"])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Act
        response = api_client.post(
            "/api/payroll/sales-revenues/upload/",
            {
                "file": file_stream,
                "month": "11/2025",
            },
            format="multipart",
        )

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["total_rows"] == 1
        assert response.data["successful"] == 1
        assert response.data["failed"] == 0
        assert SalesRevenue.objects.filter(employee=sales_employee, month=date(2025, 11, 1)).exists()

    def test_upload_updates_existing_record(self, api_client, superuser, sales_revenue):
        """Test upload updates existing record."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Mã nhân viên", "Họ tên", "Doanh Số", "Số giao dịch", "Thời gian"])
        ws.append(["Ký tự (20)", "Ký tự (100)", "Số (20)", "Số (10)", "MM/YYYY"])
        ws.append([sales_revenue.employee.code, sales_revenue.employee.username, 200000000, 20, "11/2025"])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Act
        response = api_client.post(
            "/api/payroll/sales-revenues/upload/",
            {
                "file": file_stream,
                "month": "11/2025",
            },
            format="multipart",
        )

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["successful"] == 1
        sales_revenue.refresh_from_db()
        assert sales_revenue.revenue == 200000000
        assert sales_revenue.transaction_count == 20

    def test_upload_invalid_employee_code(self, api_client, superuser):
        """Test upload with invalid employee code."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Mã nhân viên", "Họ tên", "Doanh Số", "Số giao dịch", "Thời gian"])
        ws.append(["Ký tự (20)", "Ký tự (100)", "Số (20)", "Số (10)", "MM/YYYY"])
        ws.append(["INVALID", "Test User", 100000000, 10, "11/2025"])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Act
        response = api_client.post(
            "/api/payroll/sales-revenues/upload/",
            {
                "file": file_stream,
                "month": "11/2025",
            },
            format="multipart",
        )

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["failed"] == 1
        assert len(response.data["errors"]) == 1
        assert "Employee not found" in response.data["errors"][0]["error"]

    def test_upload_month_mismatch(self, api_client, superuser, sales_employee):
        """Test upload with month mismatch."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Mã nhân viên", "Họ tên", "Doanh Số", "Số giao dịch", "Thời gian"])
        ws.append(["Ký tự (20)", "Ký tự (100)", "Số (20)", "Số (10)", "MM/YYYY"])
        ws.append([sales_employee.code, sales_employee.username, 100000000, 10, "12/2025"])

        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        # Act
        response = api_client.post(
            "/api/payroll/sales-revenues/upload/",
            {
                "file": file_stream,
                "month": "11/2025",
            },
            format="multipart",
        )

        # Assert
        assert response.status_code == status.HTTP_200_OK
        assert response.data["failed"] == 1
        assert "does not match target month" in response.data["errors"][0]["error"]

    def test_export_sales_revenue(self, api_client, superuser, sales_revenue):
        """Test exporting sales revenue to Excel."""
        # Arrange
        api_client.force_authenticate(user=superuser)

        # Act
        response = api_client.get("/api/payroll/sales-revenues/export/")

        # Assert
        assert response.status_code in [status.HTTP_200_OK, status.HTTP_202_ACCEPTED]
